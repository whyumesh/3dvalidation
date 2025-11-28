import pandas as pd
import os
from jinja2 import Environment, FileSystemLoader
import win32com.client as win32
from datetime import datetime as dt
from openpyxl import load_workbook
import glob

# Load Jinja2 Template for email
try:
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template("email_template_Division.html")
    print("✅ Email template loaded successfully")
except Exception as e:
    print(f"❌ Error loading email template: {e}")
    print("Please ensure 'email_template_Division.html' exists in the current directory")
    exit(1)

# Get current date
z = dt.today()
current_date = z.date()

# Find the most recent Division folders
def find_latest_folder(pattern):
    """Find the most recently created folder matching the pattern"""
    folders = glob.glob(pattern)
    if not folders:
        return None
    return max(folders, key=os.path.getctime)

# Get the current directory
current_dir = os.path.dirname(os.path.abspath(__file__))
if not current_dir:
    current_dir = os.getcwd()

print(f"📂 Working directory: {current_dir}")

# Locate the generated folders
consolidated_folder = find_latest_folder(os.path.join(current_dir, "Division_Consolidated_Files_*"))
reports_folder = find_latest_folder(os.path.join(current_dir, "Division_Reports_*"))

if not consolidated_folder:
    print("❌ Error: Could not find Division_Consolidated_Files folder. Please run create_division_consolidated_files.py first.")
    exit(1)

if not reports_folder:
    print("❌ Error: Could not find Division_Reports folder. Please run create_division_hierarchical_reports.py first.")
    exit(1)

print(f"✅ Found consolidated files folder: {consolidated_folder}")
print(f"✅ Found reports folder: {reports_folder}")

# Debug: List files in consolidated folder
print(f"\n🔍 Files in consolidated folder:")
if consolidated_folder and os.path.exists(consolidated_folder):
    files = os.listdir(consolidated_folder)
    print(f"   Total files: {len(files)}")
    for f in files[:5]:  # Show first 5 files
        print(f"   - {f}")
else:
    print("   Folder not accessible!")

# Debug: List files in reports folder
print(f"\n🔍 Files in reports folder:")
if reports_folder and os.path.exists(reports_folder):
    files = os.listdir(reports_folder)
    print(f"   Total files: {len(files)}")
    for f in files[:5]:  # Show first 5 files
        print(f"   - {f}")
else:
    print("   Folder not accessible!")

# Read Division Email Mapping file
print("📖 Reading Division Email Mapping file...")

# Try multiple possible file names (prioritizes actual file over sample)
# This ensures:
# - If colleague has 'division_emails.xlsx', it will use that (preferred)
# - If only 'division_emails_sample.xlsx' exists, it will use that as fallback
possible_files = ['division_emails.xlsx', 'division_emails_sample.xlsx']
division_emails_df = None
file_used = None

for filename in possible_files:
    if os.path.exists(filename):
        try:
            division_emails_df = pd.read_excel(filename)
            file_used = filename
            if filename == 'division_emails_sample.xlsx':
                print(f"✅ Successfully loaded division email mapping from: {filename} (sample file)")
                print(f"   ℹ️  Note: Using sample file. If you have 'division_emails.xlsx', it will be used instead.")
            else:
                print(f"✅ Successfully loaded division email mapping from: {filename}")
            print(f"📋 Columns in file: {list(division_emails_df.columns)}")
            break
        except Exception as e:
            print(f"⚠️ Warning: Could not read {filename}: {e}")
            continue

if division_emails_df is None:
    print(f"❌ Error: Could not find division email mapping file!")
    print(f"   Tried looking for: {', '.join(possible_files)}")
    print(f"   Please ensure one of these files exists with columns:")
    print(f"   'Affiliate', 'Division Code', 'Division Name', 'Email id'")
    print(f"\n   💡 Tip: The code will automatically use 'division_emails.xlsx' if available,")
    print(f"      otherwise it will fall back to 'division_emails_sample.xlsx'")
    exit(1)

# Map column names flexibly (handle different column name variations)
column_mapping = {}
available_columns = [col.lower() for col in division_emails_df.columns]

# Map Division Code column (try multiple variations)
if 'division code' in available_columns:
    div_code_col = division_emails_df.columns[available_columns.index('division code')]
    column_mapping['Division Code'] = div_code_col
elif 'division' in available_columns:
    div_code_col = division_emails_df.columns[available_columns.index('division')]
    column_mapping['Division Code'] = div_code_col
    division_emails_df['Division Code'] = division_emails_df[div_code_col]
else:
    print(f"❌ Error: Could not find Division Code column")
    print(f"   Available columns: {list(division_emails_df.columns)}")
    exit(1)

# Map Email id column (try multiple variations)
if 'email id' in available_columns:
    email_col = division_emails_df.columns[available_columns.index('email id')]
    column_mapping['Email id'] = email_col
elif 'email' in available_columns:
    email_col = division_emails_df.columns[available_columns.index('email')]
    column_mapping['Email id'] = email_col
    division_emails_df['Email id'] = division_emails_df[email_col]
else:
    print(f"❌ Error: Could not find Email id column")
    print(f"   Available columns: {list(division_emails_df.columns)}")
    exit(1)

print(f"✅ Column mapping:")
print(f"   Division Code: '{column_mapping['Division Code']}'")
print(f"   Email id: '{column_mapping['Email id']}'")

# Note: Affiliate and Division Name will be retrieved from the main data file

# Read ZBM Automation Email file to get division details
print("📖 Reading ZBM Automation Email 2410252.xlsx...")
df = pd.read_excel('ZBM Automation Email 2410252.xlsx')

# Get unique Divisions with their details
divisions = df.groupby('TBM Division').agg({
    'AFFILIATE': 'first',
    'DIV_NAME': 'first'
}).reset_index().sort_values('TBM Division')

print(f"📋 Found {len(divisions)} unique Divisions to process")

# Convert Division Code to string for matching (handle both string and numeric codes)
# Do this once before the loop
try:
    division_emails_df['Division Code'] = division_emails_df['Division Code'].astype(str).str.strip()
    # Also convert Email id to string and clean
    division_emails_df['Email id'] = division_emails_df['Email id'].astype(str).str.strip()
    print("✅ Division Code and Email id columns converted to string for matching")
except Exception as e:
    print(f"⚠️ Warning: Could not convert columns: {e}")

# Initialize Outlook
try:
    outlook = win32.Dispatch("Outlook.Application")
    print("✅ Outlook initialized successfully")
except Exception as e:
    print(f"❌ Error initializing Outlook: {e}")
    print("Please ensure Outlook is installed and configured")
    exit(1)

# Get the sender account (EPD_SFA@abbott.com)
sender_account = None
try:
    namespace = outlook.GetNamespace("MAPI")
    accounts = namespace.Accounts
    for account in accounts:
        if 'EPD_SFA@abbott.com' in str(account) or 'epd_sfa' in str(account).lower():
            sender_account = account
            print(f"✅ Found sender account: {account}")
            break
    
    if sender_account is None:
        print("⚠️ Warning: Could not find EPD_SFA@abbott.com account")
        print("   Available accounts:")
        for account in accounts:
            print(f"      - {account}")
        print("   Will attempt to use default account or SentOnBehalfOfName")
except Exception as e:
    print(f"⚠️ Warning: Could not access Outlook accounts: {e}")
    print("   Will attempt to use SentOnBehalfOfName instead")

# Create output directory for sent email logs
output_dir = os.path.dirname(os.path.abspath(__file__))
email_log_folder = os.path.join(output_dir, f'Division_Email_Logs_{current_date}')
os.makedirs(email_log_folder, exist_ok=True)

def read_summary_report(div_code, div_name):
    """Read the summary report Excel file for a Division and extract data as HTML table with proper formatting"""
    try:
        # Find the summary report file
        pattern = os.path.join(reports_folder, f"Division_Summary_{div_code}_*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            print(f"   ⚠️ Warning: No summary report found for Division {div_code}")
            print(f"      Searched pattern: {pattern}")
            return None
        
        report_file = os.path.abspath(files[0])
        
        # Verify file exists
        if not os.path.exists(report_file):
            print(f"   ❌ Summary report file does not exist: {report_file}")
            return None
        
        print(f"   📊 Reading summary report: {os.path.basename(report_file)}")
        
        # Read the Excel file
        wb = load_workbook(report_file)
        ws = wb.active  # Use active sheet
        
        # Find header row (looking for multiple indicators to be more robust)
        header_row = None
        for row_idx in range(1, 15):
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    # Look for multiple header indicators
                    if any(keyword in cell_str for keyword in ['Affiliate', 'ZBM Code', 'ZBM Name', 'Division', 'Requests Raised', 'Unique TBMs']):
                        header_row = row_idx
                        break
            if header_row:
                break

        if not header_row:
            print(f"   ⚠️ Warning: Could not find header row in summary report")
            return None

        # Find the first and last non-empty header columns
        first_col = None
        last_col = None
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            header_val = cell.value
            # Check if cell has value or is part of merged cell
            has_value = False
            if header_val is not None and str(header_val).strip() != "":
                has_value = True
            else:
                # Check if this cell is part of a merged range
                for merged_range in ws.merged_cells.ranges:
                    if col_idx >= merged_range.min_col and col_idx <= merged_range.max_col:
                        if header_row >= merged_range.min_row and header_row <= merged_range.max_row:
                            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            if top_left_cell.value is not None and str(top_left_cell.value).strip() != "":
                                has_value = True
                                break
            
            if has_value:
                if first_col is None:
                    first_col = col_idx
                last_col = col_idx

        if first_col is None or last_col is None:
            print(f"   ⚠️ Warning: Could not find valid header columns in summary report")
            return None

        print(f"   📊 Found header row {header_row}, columns {first_col} to {last_col}")

        # Read ALL headers from first_col to last_col
        headers = []
        header_colors = []
        
        def get_header_value(row, col):
            """Get header value handling merged cells"""
            cell = ws.cell(row=row, column=col)
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if col >= merged_range.min_col and col <= merged_range.max_col:
                    if row >= merged_range.min_row and row <= merged_range.max_row:
                        # Get the top-left cell of the merged range
                        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        return top_left_cell.value
            return cell.value
        
        for col_idx in range(first_col, last_col + 1):
            header_val = get_header_value(header_row, col_idx)
            if header_val is None:
                header_val = ""  # Use empty string for None values
            header_str = str(header_val).strip()
            headers.append(header_str)
            
            # Get background color from the actual cell
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.fill.start_color and cell.fill.start_color.rgb:
                rgb_value = cell.fill.start_color.rgb
                if isinstance(rgb_value, str):
                    header_colors.append(rgb_value)
                else:
                    header_colors.append(None)
            else:
                header_colors.append(None)
        
        print(f"   📊 Found {len(headers)} columns: {headers[:5]}..." if len(headers) > 5 else f"   📊 Found {len(headers)} columns: {headers}")

        # Build merged cells map
        merged_cells_info = {}
        for merged_range in ws.merged_cells.ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            
            # Store merge dimensions for top-left cell
            if min_row >= header_row and min_row <= ws.max_row:
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if r != min_row or c != min_col:
                            # Mark as merged cell (not top-left)
                            merged_cells_info[(r, c)] = None
        
        # Read all data rows from Excel
        rows_data = []
        empty_row_count = 0
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            has_any_value = False
            is_total_row = False
            
            # Check if this is a "Total" row by looking at first column
            first_cell_value = ws.cell(row=row_idx, column=first_col).value
            if first_cell_value and str(first_cell_value).strip().lower() == 'total':
                is_total_row = True
            
            # Check if row has any value across ALL columns
            for col_idx in range(first_col, last_col + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_any_value = True
                    break
            
            if has_any_value:
                rows_data.append({
                    'row_idx': row_idx,
                    'is_total': is_total_row
                })
                empty_row_count = 0
            else:
                empty_row_count += 1
                if empty_row_count >= 2:
                    break
        
        # Build HTML table with matching Excel formatting
        html_table = '<table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse; font-family: Arial, sans-serif; font-size: 11px;">\n'
        
        # Add header row
        html_table += '  <thead>\n    <tr style="background-color: #D3D3D3; font-weight: bold; text-align: center;">\n'
        for i, header in enumerate(headers):
            bg_color = ""
            if i < len(header_colors) and header_colors[i] and isinstance(header_colors[i], str):
                # Convert hex color if needed
                hex_color = header_colors[i]
                if hex_color.startswith('FF'):
                    hex_color = '#' + hex_color[2:]
                elif len(hex_color) == 8 and hex_color.startswith('FF'):
                    hex_color = '#' + hex_color[2:]
                bg_color = f' background-color: {hex_color};'
            html_table += f'      <th style="{bg_color} padding: 8px; border: 1px solid #000;">{header}</th>\n'
        html_table += '    </tr>\n  </thead>\n'
        
        # Add data rows with merged cell handling
        html_table += '  <tbody>\n'
        
        for row_info in rows_data:
            row_idx = row_info['row_idx']
            is_total = row_info['is_total']
            
            row_style = 'font-weight: bold; background-color: #E6E6E6;' if is_total else ''
            html_table += f'    <tr style="{row_style}">\n'
            
            # Iterate through ALL columns from first_col to last_col
            for col_idx in range(first_col, last_col + 1):
                # Skip if this cell is part of a merge (not the top-left cell)
                if (row_idx, col_idx) in merged_cells_info:
                    continue
                
                # Check if this cell starts a merge
                colspan = 1
                rowspan = 1
                
                for merged_range in ws.merged_cells.ranges:
                    if row_idx == merged_range.min_row and col_idx == merged_range.min_col:
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        colspan = merged_range.max_col - merged_range.min_col + 1
                        break
                
                # Get cell value directly from Excel
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None or pd.isna(cell_value):
                    value = '-' if not is_total_row else ''  # Empty for total row empty cells
                else:
                    value = str(cell_value).strip()
                
                # Add merge attributes if needed
                merge_attr = ''
                if rowspan > 1:
                    merge_attr += f' rowspan="{rowspan}"'
                if colspan > 1:
                    merge_attr += f' colspan="{colspan}"'
                
                html_table += f'      <td style="padding: 5px; border: 1px solid #000; text-align: center;"{merge_attr}>{value}</td>\n'
            
            html_table += '    </tr>\n'
        
        html_table += '  </tbody>\n</table>'
        
        return html_table
        
    except Exception as e:
        print(f"   ❌ Error reading summary report for Division {div_code}: {e}")
        import traceback
        traceback.print_exc()
        return None

# Removed ABM email function - using only specified CC emails from requirements

# Process each Division and send emails
email_count = 0
total_recipients = 0

for _, div_row in divisions.iterrows():
    div_code = div_row['TBM Division']
    affiliate = div_row['AFFILIATE']
    div_name = div_row['DIV_NAME']
    
    print(f"\n🔄 Processing Division: {div_code} - {affiliate} - {div_name}")
    
    # Find ALL email addresses for this division code
    # Convert div_code to string for matching
    div_code_str = str(div_code).strip()
    div_email_rows = division_emails_df[division_emails_df['Division Code'] == div_code_str]
    
    if div_email_rows.empty:
        print(f"   ⚠️ No emails found for Division Code {div_code}")
        continue
    
    # Get all valid email addresses for this division
    all_emails = []
    for _, email_row in div_email_rows.iterrows():
        email_addr = email_row['Email id']
        # Skip if no valid email
        if email_addr and str(email_addr) not in ['0', '0.0', '']:
            all_emails.append(str(email_addr).strip())
    
    if not all_emails:
        print(f"   ⚠️ No valid email addresses found for Division Code {div_code}")
        continue
    
    print(f"   📧 Found {len(all_emails)} recipient(s): {', '.join(all_emails)}")
    
    # Find consolidated file for this Division
    safe_div_name = str(div_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
    consolidated_pattern = os.path.join(consolidated_folder, f"Division_Consolidated_{div_code}_*.xlsx")
    consolidated_files = glob.glob(consolidated_pattern)
    
    if not consolidated_files:
        print(f"   ⚠️ No consolidated file found for Division {div_code}")
        print(f"      Searched pattern: {consolidated_pattern}")
        continue
    
    consolidated_file = consolidated_files[0]
    
    # Convert to absolute path (Outlook requires absolute paths)
    consolidated_file = os.path.abspath(consolidated_file)
    
    # Verify file exists
    if not os.path.exists(consolidated_file):
        print(f"   ❌ File does not exist: {consolidated_file}")
        continue
    
    print(f"   📎 Attaching: {os.path.basename(consolidated_file)}")
    
    # Read summary report data
    summary_html = read_summary_report(div_code, div_name)
    
    if not summary_html:
        print(f"   ⚠️ No summary report data found for Division {div_code}")
        continue
    
    # Build CC list based on affiliate - ONLY using specified emails from requirements
    cc_list = []
    
    # Add affiliate-specific emails ONLY
    if affiliate == 'AIL':
        cc_list = ['ishan.mithbavkar@abbott.com', 'ashwini.suryavanshi@abbott.com', 'sandesh.bhoir@abbott.com']
    elif affiliate == 'APC':
        cc_list = ['jenita.nadar@abbott.com', 'ashwini.suryavanshi@abbott.com', 'sandesh.bhoir@abbott.com']
    elif affiliate == 'ASC':
        cc_list = ['sandesh.bhoir@abbott.com', 'ashwini.suryavanshi@abbott.com']
    else:
        # For any other affiliate, default to sandesh.bhoir@abbott.com
        cc_list = ['sandesh.bhoir@abbott.com']
    
    # Remove duplicates and join
    cc_list = list(set(cc_list))
    final_cc = '; '.join(cc_list)
    
    # Create email with ALL recipients in TO field
    try:
        mail = outlook.CreateItem(0)
        
        # Set sender account BEFORE setting other properties
        # Method 1: Try SendUsingAccount (most reliable if account is configured in Outlook)
        sender_set = False
        if sender_account:
            try:
                mail.SendUsingAccount = sender_account
                print(f"   ✅ Using sender account: {sender_account}")
                sender_set = True
            except Exception as account_error:
                print(f"   ⚠️ Warning: Could not set SendUsingAccount: {account_error}")
                print(f"      Error details: {str(account_error)}")
        
        # Method 2: Fallback to SentOnBehalfOfName (requires 'Send As' permissions)
        if not sender_set:
            try:
                mail.SentOnBehalfOfName = 'EPD_SFA@abbott.com'
                print(f"   ✅ Using SentOnBehalfOfName: EPD_SFA@abbott.com")
                sender_set = True
            except Exception as behalf_error:
                print(f"   ⚠️ Warning: Could not set SentOnBehalfOfName: {behalf_error}")
                print(f"      Error details: {str(behalf_error)}")
        
        if not sender_set:
            print(f"   ⚠️ WARNING: Could not set sender account. Email will use default Outlook account.")
            print(f"   ⚠️ To fix this:")
            print(f"      1. Ensure EPD_SFA@abbott.com is added as an account in Outlook")
            print(f"      2. OR ensure you have 'Send As' permissions for EPD_SFA@abbott.com")
            print(f"      3. Check Outlook account settings and permissions")
        
        # Add ALL email addresses to TO field (semicolon-separated)
        mail.To = '; '.join(all_emails)
        
        # Add CC recipients
        if final_cc:
            mail.CC = final_cc
            print(f"   📧 CC: {final_cc}")
        
        # Set BCC
        mail.BCC = 'vaibhav.nalawade@abbott.com;kranti.vengurlekar@abbott.com'
        
        # Set subject
        mail.Subject = f"{div_name}: Sample Direct Dispatch to Doctors - Request Status as of {current_date}"
        
        # Render email body with summary table
        try:
            mail.HTMLBody = template.render(
                division_name=div_name,
                division_code=div_code,
                affiliate=affiliate,
                current_date=current_date,
                summary_table=summary_html
            )
        except Exception as template_error:
            print(f"   ❌ Error rendering email template: {template_error}")
            continue
        
        # Attach consolidated file - AFTER setting body
        try:
            mail.Attachments.Add(consolidated_file)
            print(f"   ✅ Attachment added successfully")
        except Exception as attach_error:
            print(f"   ❌ Error attaching file: {attach_error}")
            print(f"      File path: {consolidated_file}")
            continue
        
        # Display email
        mail.Display()
        
        email_count += 1
        total_recipients += len(all_emails)
        print(f"   ✅ Email displayed successfully for {len(all_emails)} recipient(s)")
        
        # Log the sent email
        with open(os.path.join(email_log_folder, 'email_log.txt'), 'a') as log:
            log.write(f"{dt.now()} - Displayed email for Division {div_code} ({div_name})\n")
            log.write(f"   TO: {'; '.join(all_emails)}\n")
            log.write(f"   CC: {final_cc}\n")
            log.write(f"   BCC: vaibhav.nalawade@abbott.com;kranti.vengurlekar@abbott.com\n\n")
        
    except Exception as e:
        print(f"   ❌ Error creating email for Division {div_code}: {e}")
        import traceback
        traceback.print_exc()
        continue

print(f"\n🎉 Email automation completed!")
print(f"📊 Total emails displayed: {email_count} for {total_recipients} total recipients")
print(f"📁 Email logs saved in: {email_log_folder}")
