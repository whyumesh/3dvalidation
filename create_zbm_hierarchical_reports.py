import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy as copy_style
import warnings

# Suppress FutureWarning for groupby operations
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

def create_zbm_hierarchical_reports():
    """
    Create separate ZBM reports showing ABM hierarchy with perfect tallies
    Each ZBM gets a report showing all ABMs under them
    """
    
    print("🔄 Starting ZBM Hierarchical Reports Creation...")
    
    # Read master tracker data from Excel file
    print("📖 Reading Demo File 1.xlsx...")
    try:
        df = pd.read_excel('Demo File 1.xlsx')
        print(f"✅ Successfully loaded {len(df)} records")
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return
    
    print(f"📋 Columns in file: {list(df.columns)}")
    
    # Basic data preparation
    print("🧹 Preparing data...")
    
    # Ensure required columns exist
    required_columns = ['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID',
                        'ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID',
                        'TBM HQ', 'TBM EMAIL_ID',
                        'Doctor: Customer Code', 'Assigned Request Ids', 'Request Status', 'Rto Reason']
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns: {missing}")
        return

    print(f"📊 Total rows in file: {len(df)}")
    print(f"📊 Unique Request IDs in raw data: {df['Assigned Request Ids'].nunique()}")
    print(f"📊 Unique ZBM codes in raw data: {df['ZBM Terr Code'].nunique()}")

    # Compute Final Answer per unique request id using rules from logic.xlsx
    print("🧠 Computing final status per unique Request Id using rules...")
    try:
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        def normalize(text):
            return str(text).strip().casefold()

        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']

        # Group statuses by request id from master data
        grouped = df.groupby('Assigned Request Ids')['Request Status'].apply(list).reset_index()

        def get_final_answer(status_list):
            key = tuple(sorted(set(normalize(s) for s in status_list)))
            return rules.get(key, '❌ No matching rule')

        grouped['Request Status'] = grouped['Request Status'].apply(lambda lst: sorted(set(lst), key=str))
        grouped['Final Answer'] = grouped['Request Status'].apply(get_final_answer)

        # Merge Final Answer back to main dataframe
        df = df.merge(grouped[['Assigned Request Ids', 'Final Answer']], on='Assigned Request Ids', how='left')
        
        # Check for unmapped requests
        unmapped_count = (df['Final Answer'] == '❌ No matching rule').sum()
        if unmapped_count > 0:
            print(f"⚠️ WARNING: {unmapped_count} rows have no matching rule in logic.xlsx")
            print(f"   Unique Request IDs with no rule: {df[df['Final Answer'] == '❌ No matching rule']['Assigned Request Ids'].nunique()}")
            
    except Exception as e:
        print(f"❌ Error computing final status from logic.xlsx: {e}")
        return
    
    # Deduplicate at Request ID + ZBM + ABM level to get correct counts
    print("🔧 Deduplicating data at Request ID + ZBM + ABM level...")
    
    # Store original data for validation
    original_request_count = df['Assigned Request Ids'].nunique()
    
    # Deduplicate: Each unique (Request ID + ZBM + ABM) combination should appear once
    df_dedup = df.groupby(['Assigned Request Ids', 'ZBM Terr Code', 'ABM Terr Code']).agg({
        'ZBM Name': 'first',
        'ZBM EMAIL_ID': 'first',
        'ABM Name': 'first',
        'ABM EMAIL_ID': 'first',
        'TBM HQ': 'first',
        'TBM EMAIL_ID': 'first',
        'Doctor: Customer Code': 'first',
        'Final Answer': 'first',
        'Rto Reason': 'first',
        'ABM HQ': 'first' if 'ABM HQ' in df.columns else lambda x: None
    }).reset_index()
    
    print(f"📊 Deduplicated from {len(df)} rows to {len(df_dedup)} unique (Request ID + ZBM + ABM) combinations")
    print(f"📊 Unique Request IDs after dedup: {df_dedup['Assigned Request Ids'].nunique()}")
    
    # Get unique ZBMs using mode (most frequent) for name/email
    zbms = df_dedup.groupby('ZBM Terr Code').agg({
        'ZBM Name': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
        'ZBM EMAIL_ID': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0]
    }).reset_index().sort_values('ZBM Terr Code')
    
    print(f"📋 Found {len(zbms)} unique ZBMs")
    
    # Debug: Show first few ZBMs and their ABMs
    print("\n🔍 ZBM-ABM Mapping (first 5):")
    for idx, (_, zbm_row) in enumerate(zbms.head(5).iterrows()):
        zbm_code = zbm_row['ZBM Terr Code']
        zbm_name = zbm_row['ZBM Name']
        zbm_data_temp = df_dedup[df_dedup['ZBM Terr Code'] == zbm_code]
        abms_temp = zbm_data_temp[['ABM Terr Code', 'ABM Name']].drop_duplicates()
        requests_temp = zbm_data_temp['Assigned Request Ids'].nunique()
        print(f"   {idx+1}. {zbm_code} ({zbm_name}): {len(abms_temp)} ABMs, {requests_temp} requests")
    
    # Create output directory
    timestamp = datetime.now().strftime('%Y%m%d')
    output_dir = f"ZBM_Reports_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Created output directory: {output_dir}")
    
    # Process each ZBM
    file_count = 0
    total_validation_errors = 0
    
    for _, zbm_row in zbms.iterrows():
        zbm_code = zbm_row['ZBM Terr Code']
        zbm_name = zbm_row['ZBM Name']
        zbm_email = zbm_row['ZBM EMAIL_ID']
        
        print(f"\n🔄 Processing ZBM: {zbm_code} - {zbm_name}")
        
        # Filter data for this ZBM (using deduplicated data)
        zbm_data = df_dedup[df_dedup['ZBM Terr Code'] == zbm_code].copy()
        
        if len(zbm_data) == 0:
            print(f"⚠️ No data found for ZBM: {zbm_code}")
            continue
        
        # Get unique ABMs under this ZBM
        abms = zbm_data.groupby(['ABM Terr Code', 'ABM Name']).agg({
            'ABM EMAIL_ID': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
            'TBM HQ': 'first',
            'ABM HQ': 'first' if 'ABM HQ' in zbm_data.columns else lambda x: None
        }).reset_index()
        
        abms = abms.sort_values('ABM Terr Code')
        print(f"   📊 Found {len(abms)} ABMs under this ZBM")
        
        # Create summary data for this ZBM
        summary_data = []
        
        for _, abm_row in abms.iterrows():
            abm_code = abm_row['ABM Terr Code']
            abm_name = abm_row['ABM Name']
            abm_email = abm_row['ABM EMAIL_ID']
            tbm_hq = abm_row['TBM HQ']
            
            # Filter data for this specific ABM (using deduplicated data)
            abm_data = zbm_data[(zbm_data['ABM Terr Code'] == abm_code) & 
                               (zbm_data['ABM Name'] == abm_name)].copy()
            
            # Calculate all metrics using deduplicated data
            unique_tbms = abm_data['TBM EMAIL_ID'].nunique()
            unique_hcps = abm_data['Doctor: Customer Code'].nunique()
            unique_requests = len(abm_data)
            
            # === SECTION A: Request Cancelled Out of Stock ===
            # Final Answer: Out of stock, On hold, Not permitted
            ho_statuses = ['Out of stock', 'On hold', 'Not permitted']
            request_cancelled_out_of_stock = (abm_data['Final Answer'].isin(ho_statuses)).sum()
            
            # === SECTION B: Action Pending at HO ===
            # Final Answer: Request Raised, Action pending / In Process At HO
            pending_statuses = ['Request Raised', 'Action pending / In Process At HO']
            action_pending_at_ho = (abm_data['Final Answer'].isin(pending_statuses)).sum()
            
            # === SECTION D: Pending for Invoicing ===
            # Final Answer: Action pending / In Process At Hub
            hub_pending_statuses = ['Action pending / In Process At Hub']
            pending_for_invoicing = (abm_data['Final Answer'].isin(hub_pending_statuses)).sum()
            
            # === SECTION E: Pending for Dispatch ===
            # Final Answer: Dispatch Pending
            dispatch_pending_statuses = ['Dispatch  Pending', 'Dispatch Pending']
            pending_for_dispatch = (abm_data['Final Answer'].isin(dispatch_pending_statuses)).sum()
            
            # === SECTION G: Delivered ===
            # Final Answer: Delivered
            delivered_statuses = ['Delivered']
            delivered = (abm_data['Final Answer'].isin(delivered_statuses)).sum()
            
            # === SECTION H: Dispatched & In Transit ===
            # Final Answer: Dispatched & In Transit
            transit_statuses = ['Dispatched & In Transit']
            dispatched_in_transit = (abm_data['Final Answer'].isin(transit_statuses)).sum()
            
            # === SECTION I: RTO (Return to Origin) ===
            # RTO Calculation - ONLY count requests with "Return" Final Answer
            # Priority-Based Reason Assignment: 1) Incomplete Address, 2) Doctor Refused, 3) Doctor Non Contactable
            
            # ONLY count requests with "Return" Final Answer as RTO
            has_return_status = abm_data['Final Answer'] == 'Return'
            rto_total = has_return_status.sum()
            
            # RTO Reasons - Check Rto Reason column (only for Return requests)
            rto_col = abm_data['Rto Reason'].astype(str).str.strip().str.lower()
            has_incomplete_address = rto_col.str.contains('incomplete address', na=False, regex=False)
            has_refused_to_accept = rto_col.str.contains('refused to accept', na=False, regex=False)
            has_non_contactable = rto_col.str.contains('non contactable', na=False, regex=False)
            
            # Any RTO reason present
            has_any_rto_reason = has_incomplete_address | has_non_contactable | has_refused_to_accept
            
            # Assign each RTO request to ONE category based on priority
            # Only count requests that have Return status
            incomplete_address = (has_return_status & has_incomplete_address).sum()
            doctor_refused_to_accept = (has_return_status & ~has_incomplete_address & has_refused_to_accept).sum()
            doctor_non_contactable = (has_return_status & ~has_incomplete_address & ~has_non_contactable & has_refused_to_accept).sum()
            
            # Handle Return status without RTO reason - add to Non Contactable as catch-all
            return_no_reason = (has_return_status & ~has_any_rto_reason).sum()
            if return_no_reason > 0:
                doctor_non_contactable += return_no_reason
            
            # Validate RTO breakdown
            rto_reasons_sum = incomplete_address + doctor_non_contactable + doctor_refused_to_accept
            if rto_reasons_sum != rto_total:
                print(f"      ⚠️ RTO Breakdown mismatch for ABM {abm_code}:")
                print(f"         RTO Total: {rto_total}, Reasons Sum: {rto_reasons_sum}")
                print(f"         Incomplete: {incomplete_address}, Refused: {doctor_refused_to_accept}, Non-contactable: {doctor_non_contactable}")
            
            # === CALCULATED FIELDS ===
            # F = Requests Dispatched = G + H + I
            requests_dispatched = delivered + dispatched_in_transit + rto_total
            
            # C = Sent to HUB = D + E + F
            sent_to_hub = pending_for_invoicing + pending_for_dispatch + requests_dispatched
            
            # Total = Requests Raised = A + B + C
            requests_raised_calc = request_cancelled_out_of_stock + action_pending_at_ho + sent_to_hub
            
            # Hold Delivery (not used in current logic)
            hold_delivery = 0
            
            # Check for unmapped requests
            all_mapped_statuses = ho_statuses + pending_statuses + hub_pending_statuses + dispatch_pending_statuses + delivered_statuses + transit_statuses + ['Return']
            mapped = abm_data['Final Answer'].isin(all_mapped_statuses)
            unmapped_count = (~mapped).sum()
            
            if unmapped_count > 0:
                print(f"      ⚠️ {unmapped_count} unmapped requests for ABM {abm_code}")
                unmapped_statuses = abm_data[~mapped]['Final Answer'].value_counts().to_dict()
                print(f"         Unmapped statuses: {unmapped_statuses}")
            
            # Verify tally
            if requests_raised_calc != unique_requests:
                print(f"      ❌ TALLY MISMATCH for ABM {abm_code}:")
                print(f"         Calculated: {requests_raised_calc}, Actual: {unique_requests}, Diff: {unique_requests - requests_raised_calc}")
                print(f"         A={request_cancelled_out_of_stock}, B={action_pending_at_ho}, C={sent_to_hub}")
                print(f"         D={pending_for_invoicing}, E={pending_for_dispatch}, F={requests_dispatched}")
                print(f"         G={delivered}, H={dispatched_in_transit}, I={rto_total}")
                total_validation_errors += 1
            
            # Use actual unique request count to ensure accuracy
            requests_raised = unique_requests
            
            # Create Area Name
            if 'ABM HQ' in abm_row and pd.notna(abm_row['ABM HQ']):
                abm_hq = abm_row['ABM HQ']
            else:
                abm_hq = tbm_hq
            area_name = f"{abm_code} - {abm_hq}"
            
            summary_data.append({
                'Area Name': area_name,
                'ABM Name': abm_name,
                'Unique TBMs': unique_tbms,
                'Unique HCPs': unique_hcps,
                'Requests Raised': requests_raised,
                'Request Cancelled Out of Stock': request_cancelled_out_of_stock,
                'Action Pending at HO': action_pending_at_ho,
                'Sent to HUB': sent_to_hub,
                'Pending for Invoicing': pending_for_invoicing,
                'Pending for Dispatch': pending_for_dispatch,
                'Requests Dispatched': requests_dispatched,
                'Delivered': delivered,
                'Dispatched In Transit': dispatched_in_transit,
                'RTO': rto_total,
                'Incomplete Address': incomplete_address,
                'Doctor Non Contactable': doctor_non_contactable,
                'Doctor Refused to Accept': doctor_refused_to_accept,
                'Hold Delivery': hold_delivery
            })
        
        # Create DataFrame for this ZBM
        zbm_summary_df = pd.DataFrame(summary_data)
        
        # Validate ZBM total
        zbm_total_requests = zbm_data['Assigned Request Ids'].nunique()
        zbm_summary_total = zbm_summary_df['Requests Raised'].sum()
        
        if zbm_total_requests != zbm_summary_total:
            print(f"   ⚠️ WARNING: ZBM {zbm_code} total mismatch!")
            print(f"      Actual unique requests: {zbm_total_requests}")
            print(f"      Summary total: {zbm_summary_total}")
            print(f"      Difference: {zbm_summary_total - zbm_total_requests}")
        
        # Create Excel file for this ZBM
        create_zbm_excel_report(zbm_code, zbm_name, zbm_email, zbm_summary_df, output_dir)
        file_count += 1
    
    print(f"\n🎉 Successfully created {file_count} ZBM reports in directory: {output_dir}")
    print(f"📊 Total ZBMs processed: {file_count}")
    if total_validation_errors > 0:
        print(f"⚠️ WARNING: {total_validation_errors} ABMs had validation errors")
    else:
        print(f"✅ All tallies match perfectly!")

def create_zbm_excel_report(zbm_code, zbm_name, zbm_email, summary_df, output_dir):
    """Create Excel report for a specific ZBM with perfect formatting"""
    
    try:
        # Load template
        wb = load_workbook('zbm_summary.xlsx')
        ws = wb['ZBM']

        def get_cell_value_handling_merged(row, col):
            """Get cell value even if it's part of a merged cell"""
            cell = ws.cell(row=row, column=col)
            
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return top_left_cell.value
            
            return cell.value
        
        # Search for header row
        header_row = None
        for row_idx in range(1, 15):
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = get_cell_value_handling_merged(row_idx, col_idx)
                if cell_value and 'Area Name' in str(cell_value):
                    header_row = row_idx
                    break
            if header_row:
                break
        
        if header_row is None:
            header_row = 7
        
        data_start_row = header_row + 1
        
        # Read actual column positions from template header row
         column_mapping = {}
        for col_idx in range(1, min(30, ws.max_column + 1)):
            header_val = get_cell_value_handling_merged(header_row, col_idx)
            if header_val:
                header_str = str(header_val).strip()
                
                if 'Area Name' in header_str:
                    column_mapping['Area Name'] = col_idx
                elif 'ABM Name' in header_str:
                    column_mapping['ABM Name'] = col_idx
                elif 'Unique TBMs' in header_str or '# Unique TBMs' in header_str:
                    column_mapping['Unique TBMs'] = col_idx
                elif 'Unique HCPs' in header_str or '# Unique HCPs' in header_str:
                    column_mapping['Unique HCPs'] = col_idx
                elif 'Requests Raised' in header_str or '# Requests Raised' in header_str:
                    column_mapping['Requests Raised'] = col_idx
                elif 'Request Cancelled' in header_str or 'Out of Stock' in header_str:
                    column_mapping['Request Cancelled Out of Stock'] = col_idx
                elif 'Action pending' in header_str and 'HO' in header_str:
                    column_mapping['Action Pending at HO'] = col_idx
                elif 'Sent to HUB' in header_str:
                    column_mapping['Sent to HUB'] = col_idx
                elif 'Pending for Invoicing' in header_str:
                    column_mapping['Pending for Invoicing'] = col_idx
                elif 'Pending for Dispatch' in header_str:
                    column_mapping['Pending for Dispatch'] = col_idx
                elif 'Requests Dispatched' in header_str or '# Requests Dispatched' in header_str:
                    column_mapping['Requests Dispatched'] = col_idx
                elif header_str == 'Delivered' or 'Delivered (G)' in header_str:
                    column_mapping['Delivered'] = col_idx
                elif 'Dispatched & In Transit' in header_str or 'Dispatched In Transit' in header_str:
                    column_mapping['Dispatched In Transit'] = col_idx
                elif header_str == 'RTO' or 'RTO (I)' in header_str:
                    column_mapping['RTO'] = col_idx
                elif 'Incomplete Address' in header_str:
                    column_mapping['Incomplete Address'] = col_idx
                elif 'Doctor Non Contactable' in header_str or 'Dr. Non contactable' in header_str:
                    column_mapping['Doctor Non Contactable'] = col_idx
                elif 'Doctor Refused' in header_str or 'Refused to Accept' in header_str:
                    column_mapping['Doctor Refused to Accept'] = col_idx
                elif 'Hold Delivery' in header_str:
                    column_mapping['Hold Delivery'] = col_idx
        
        # Clear existing data rows
        max_clear_rows = max(len(summary_df) + 10, 50)
        for r in range(data_start_row, data_start_row + max_clear_rows):
            for c in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                except:
                    pass

        def copy_row_style(src_row_idx, dst_row_idx):
            """Copy formatting from source row to destination row"""
            for c in range(1, ws.max_column + 1):
                try:
                    src = ws.cell(row=src_row_idx, column=c)
                    dst = ws.cell(row=dst_row_idx, column=c)
                    
                    if src.font:
                        dst.font = copy_style(src.font)
                    if src.alignment:
                        dst.alignment = copy_style(src.alignment)
                    if src.border:
                        dst.border = copy_style(src.border)
                    if src.fill:
                        dst.fill = copy_style(src.fill)
                    dst.number_format = src.number_format
                except:
                    pass

        # Write data rows
        template_data_row = data_start_row
        for i in range(len(summary_df)):
            target_row = data_start_row + i
            copy_row_style(template_data_row, target_row)
            
            for col_name, col_idx in column_mapping.items():
                if col_name in summary_df.columns:
                    value = summary_df.iloc[i][col_name]
                    
                    try:
                        cell = ws.cell(row=target_row, column=col_idx)
                        cell.value = value
                        
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0'
                    except:
                        pass

        # Add total row
        total_row = data_start_row + len(summary_df)
        copy_row_style(template_data_row, total_row)
        
        if 'ABM Name' in column_mapping:
            try:
                cell = ws.cell(row=total_row, column=column_mapping['ABM Name'])
                cell.value = "Total"
                cell.font = Font(bold=True, name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except:
                pass
        
        # Calculate and write totals
        for col_name, col_idx in column_mapping.items():
            if col_name in summary_df.columns and col_name not in ['Area Name', 'ABM Name']:
                total_value = int(summary_df[col_name].sum())
                
                try:
                    cell = ws.cell(row=total_row, column=col_idx)
                    cell.value = total_value
                    cell.font = Font(bold=True, name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = '0'
                except:
                    pass

        # Save file
        safe_zbm_name = str(zbm_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f"ZBM_Summary_{zbm_code}_{safe_zbm_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb.save(filepath)
        print(f"   ✅ Created: {filename}")
        
    except Exception as e:
        print(f"   ❌ Error creating Excel report for {zbm_code}: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    create_zbm_hierarchical_reports()
