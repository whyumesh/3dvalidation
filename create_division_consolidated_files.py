#!/usr/bin/env python3
"""
Division Consolidated Files Generator
Creates detailed consolidated files for each TBM Division with specific columns from Sample Master Tracker
"""

import pandas as pd
import os
from datetime import datetime
import warnings

# Suppress pandas warnings
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

def create_division_consolidated_files():
    """Create consolidated files for each TBM Division with detailed data"""
    
    print("🚀 Starting Division Consolidated Files Generation...")
    
    # Read Sample Master Tracker data
    print("📖 Reading ZBM Automation Email 2410252.xlsx...")
    try:
        df = pd.read_excel('ZBM Automation Email 2410252.xlsx')
        print(f"✅ Successfully loaded {len(df)} records from ZBM Automation Email 2410252.xlsx")
    except Exception as e:
        print(f"❌ Error reading ZBM Automation Email 2410252.xlsx: {e}")
        return
    
    # Required columns for consolidated file
    required_columns = [
        'Assigned Request Ids', 'Doctor: SAP Customer Code(New)', 'Doctor: Customer Code', 
        'Doctor: Account Name', 'Item Code', 'SKU', 'Requested Quantity', 'TBM Division', 
        'AFFILIATE', 'DIV_NAME', 'Date', 'Month', 'Invoice #', 'Invoice Date', 
        'Dispatch Date', 'Delivery Date', 'Docket Number', 'Transporter Name', 
        'Request Status', 'Rto Reason', 'Input Sample Request: Created By', 'TBM HQ', 
        'ABM Name', 'ABM Terr Code', 'ZBM Name', 'ZBM Terr Code'
    ]
    
    # Check for missing columns
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns in ZBM Automation Email 2410252.xlsx: {missing}")
        print(f"📋 Available columns: {list(df.columns)}")
        return
    
    # Debug: Check RTO Reason column data
    print(f"🔍 RTO Reason column analysis:")
    rto_reason_counts = df['Rto Reason'].value_counts(dropna=False)
    print(f"   Total records: {len(df)}")
    print(f"   Non-null RTO Reasons: {df['Rto Reason'].notna().sum()}")
    print(f"   RTO Reason value counts:")
    for reason, count in rto_reason_counts.head(10).items():
        print(f"      '{reason}': {count}")
    
    # Debug: Check TBM Division column data
    print(f"🔍 TBM Division column analysis:")
    tbm_div_counts = df['TBM Division'].value_counts(dropna=False)
    print(f"   Non-null TBM Divisions: {df['TBM Division'].notna().sum()}")
    print(f"   TBM Division value counts (top 10):")
    for div, count in tbm_div_counts.head(10).items():
        print(f"      '{div}': {count}")
    
    # Remove rows where key fields are null or empty
    df = df.dropna(subset=['TBM Division', 'ABM Terr Code', 'ABM Name'])
    df = df[df['TBM Division'].astype(str).str.strip() != '']
    df = df[df['ABM Terr Code'].astype(str).str.strip() != '']
    
    # Standardize Division data to prevent duplicates
    df['TBM Division'] = df['TBM Division'].astype(str).str.strip()
    df['ABM Terr Code'] = df['ABM Terr Code'].astype(str).str.strip()
    
    print(f"📊 After cleaning: {len(df)} records remaining")
    
    # Compute Final Answer per unique request id using rules from logic.xlsx
    print("🧠 Computing final status per unique Request Id using rules...")
    try:
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        def normalize(text):
            return str(text).strip().casefold()

        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']

        # Group statuses by request id from master data
        grouped = df.groupby('Assigned Request Ids')['Request Status'].apply(list).reset_index()

        def get_final_answer(status_list):
            key = tuple(sorted(set(normalize(s) for s in status_list)))
            return rules.get(key, '❌ No matching rule')

        grouped['Request Status'] = grouped['Request Status'].apply(lambda lst: sorted(set(lst), key=str))
        grouped['Final Answer'] = grouped['Request Status'].apply(get_final_answer)

        def has_action_pending(status_list):
            target = 'action pending / in process'
            return any(normalize(s) == target for s in status_list)
        grouped['Has D Pending'] = grouped['Request Status'].apply(has_action_pending)

        # Merge Final Answer back to main dataframe
        df = df.merge(grouped[['Assigned Request Ids', 'Final Answer', 'Has D Pending']], on='Assigned Request Ids', how='left')
        
        # Use Final Answer as Final Status
        df['Final Status'] = df['Final Answer']
        
        print(f"✅ Successfully computed final status for all records")
        
    except Exception as e:
        print(f"❌ Error computing final status from logic.xlsx: {e}")
        # If logic file fails, use Request Status as Final Status
        df['Final Status'] = df['Request Status']
    
    # First, let's see how many unique TBM Divisions exist
    unique_tbm_divisions = df['TBM Division'].unique()
    print(f"📋 Found {len(unique_tbm_divisions)} unique TBM Divisions")
    
    # Now let's check if same TBM Division has different AFFILIATE/DIV_NAME
    div_check = df.groupby('TBM Division')[['AFFILIATE', 'DIV_NAME']].nunique()
    duplicates = div_check[(div_check['AFFILIATE'] > 1) | (div_check['DIV_NAME'] > 1)]
    
    if len(duplicates) > 0:
        print(f"⚠️ WARNING: Found {len(duplicates)} TBM Divisions with multiple AFFILIATE/DIV_NAME:")
        for div_code in duplicates.index:
            affiliates = df[df['TBM Division'] == div_code]['AFFILIATE'].unique()
            div_names = df[df['TBM Division'] == div_code]['DIV_NAME'].unique()
            print(f"   Division {div_code} has {len(affiliates)} different affiliates and {len(div_names)} different div names")
    
    # Get unique TBM Divisions - Use ONLY TBM Division for uniqueness
    div_groups = df.groupby('TBM Division').agg({
        'AFFILIATE': 'first',  # Take the first affiliate
        'DIV_NAME': 'first'    # Take the first div name
    }).reset_index()
    
    divisions = div_groups.sort_values('TBM Division')
    
    print(f"\n📋 Creating {len(divisions)} consolidated files (one per unique TBM Division)")
    
    # Debug: Show Division list
    print(f"🔍 Unique TBM Divisions to be processed:")
    for _, div in divisions.iterrows():
        div_count = len(df[df['TBM Division'] == div['TBM Division']])
        print(f"   Division {div['TBM Division']} - {div['AFFILIATE']} - {div['DIV_NAME']} ({div_count} records)")
    
    # Create output directory
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = f"Division_Consolidated_Files_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Created output directory: {output_dir}")
    
    # Process each TBM Division
    for _, div_row in divisions.iterrows():
        div_code = div_row['TBM Division']
        affiliate = div_row['AFFILIATE']
        div_name = div_row['DIV_NAME']
        
        print(f"\n🔄 Processing TBM Division: {div_code} - {affiliate} - {div_name}")
        
        # Filter data for this TBM Division
        div_data = df[df['TBM Division'] == div_code]
        
        if len(div_data) == 0:
            print(f"⚠️ No data found for TBM Division: {div_code}")
            continue
        
        print(f"   📊 Found {len(div_data)} records for this TBM Division")
        
        # Select only the required columns for consolidated file
        consolidated_columns = [
            'Assigned Request Ids', 'Doctor: SAP Customer Code(New)', 'Doctor: Customer Code', 
            'Doctor: Account Name', 'Item Code', 'SKU', 'Requested Quantity', 'TBM Division', 
            'AFFILIATE', 'DIV_NAME', 'Date', 'Month', 'Invoice #', 'Invoice Date', 
            'Dispatch Date', 'Delivery Date', 'Docket Number', 'Transporter Name', 
            'Request Status', 'Final Status', 'Rto Reason', 'Input Sample Request: Created By', 'TBM HQ', 
            'ABM Name', 'ABM Terr Code', 'ZBM Name', 'ZBM Terr Code'
        ]
        
        # Create consolidated data for this TBM Division
        consolidated_data = div_data[consolidated_columns].copy()
        
        # Format date columns to show only date without time
        date_columns = ['Date', 'Invoice Date', 'Dispatch Date', 'Delivery Date']
        for col in date_columns:
            if col in consolidated_data.columns:
                # Convert to datetime and format as date only
                consolidated_data[col] = pd.to_datetime(consolidated_data[col], errors='coerce').dt.date
                print(f"   📅 Formatted {col} column to date-only format")
        
        # Sort by ABM Terr Code and then by Assigned Request Ids
        consolidated_data = consolidated_data.sort_values(['ABM Terr Code', 'Assigned Request Ids'])
        
        # Create filename
        safe_div_name = str(div_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f"Division_Consolidated_{div_code}_{safe_div_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Save to Excel
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                consolidated_data.to_excel(writer, sheet_name='Consolidated Data', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Consolidated Data']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add header formatting
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, name='Arial', size=10)
                header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Format date columns in Excel
                date_columns_excel = ['Date', 'Invoice Date', 'Dispatch Date', 'Delivery Date']
                for col_name in date_columns_excel:
                    if col_name in consolidated_data.columns:
                        # Find the column index
                        col_idx = consolidated_data.columns.get_loc(col_name) + 1  # +1 because Excel is 1-indexed
                        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                        
                        # Apply date format to all data cells in this column
                        for row in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
                            cell = worksheet.cell(row=row, column=col_idx)
                            if cell.value is not None:
                                cell.number_format = 'dd/mm/yyyy'  # Date format without time
            
            print(f"   ✅ Created: {filename}")
            print(f"   📊 Records in consolidated file: {len(consolidated_data)}")
            
            # Show sample of data
            print(f"   📋 Sample data (first 3 rows):")
            for idx, (_, row) in enumerate(consolidated_data.head(3).iterrows()):
                print(f"      Row {idx+1}: {row['ABM Name']} - {row['Assigned Request Ids']} - {row['Request Status']} -> {row['Final Status']} - RTO: {row['Rto Reason']}")
            
            # Debug: Check RTO Reason data in this Division's consolidated file
            rto_reason_data = consolidated_data['Rto Reason'].value_counts(dropna=False)
            print(f"   🔍 RTO Reason data in consolidated file:")
            print(f"      Non-null RTO Reasons: {consolidated_data['Rto Reason'].notna().sum()}")
            for reason, count in rto_reason_data.head(5).items():
                print(f"      '{reason}': {count}")
            
        except Exception as e:
            print(f"   ❌ Error creating consolidated file for Division {div_code}: {e}")
            continue
    
    print(f"\n🎉 Successfully created {len(divisions)} consolidated files in directory: {output_dir}")
    print(f"📁 Each file contains detailed data for that specific TBM Division only")
    print(f"📧 These files are ready to be attached to Division emails")

if __name__ == "__main__":
    create_division_consolidated_files()
