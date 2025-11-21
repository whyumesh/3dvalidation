import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy as copy_style
import warnings

# Suppress FutureWarning for groupby operations
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

def create_division_hierarchical_reports():
    """
    Create separate Division reports showing TBM Division hierarchy with perfect tallies
    Each TBM Division gets a report showing all TBMs under them
    """
    
    print("📄 Starting Division Hierarchical Reports Creation...")
    
    # Read master tracker data from Excel file
    print("📖 Reading ZBM Automation Email 2410252.xlsx...")
    try:
        df = pd.read_excel('ZBM Automation Email 2410252.xlsx')
        print(f"✅ Successfully loaded {len(df)} records")
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return
    
    print(f"📋 Columns in file: {list(df.columns)}")
    
    # Basic data preparation
    print("🧹 Preparing data...")
    
    # Find the correct column name for TBM/Created By
    tbm_created_by_col = None
    for col in df.columns:
        if 'created by' in col.lower() or 'created_by' in col.lower():
            tbm_created_by_col = col
            print(f"✅ Found TBM Created By column: '{col}'")
            break
    
    if tbm_created_by_col is None:
        print("Warning: Could not find 'Created By' column, will use 'TBM EMAIL_ID' instead")
        tbm_created_by_col = 'TBM EMAIL_ID'
    
    # Ensure required columns exist
    required_columns = ['TBM Division', 'AFFILIATE', 'DIV_NAME',
                        'ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID',
                        'ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID',
                        'Doctor: Customer Code', 'Assigned Request Ids', 'Request Status', 'Rto Reason']
    
    # Add the TBM created by column if it's different from TBM EMAIL_ID
    if tbm_created_by_col != 'TBM EMAIL_ID' and tbm_created_by_col not in required_columns:
        required_columns.append(tbm_created_by_col)
    
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns: {missing}")
        return

    print(f"📊 Total rows in file: {len(df)}")
    print(f"📊 Unique Request IDs in raw data: {df['Assigned Request Ids'].nunique()}")
    print(f"📊 Unique TBM Divisions in raw data: {df['TBM Division'].nunique()}")

    # Compute Final Answer per unique request id using rules from logic.xlsx
    print("🧠 Computing final status per unique Request Id using rules...")
    try:
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        def normalize(text):
            return str(text).strip().casefold()

        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']

        # Group statuses by request id from master data
        grouped = df.groupby('Assigned Request Ids')['Request Status'].apply(list).reset_index()

        def get_final_answer(status_list):
            key = tuple(sorted(set(normalize(s) for s in status_list)))
            return rules.get(key, '❌ No matching rule')

        grouped['Request Status'] = grouped['Request Status'].apply(lambda lst: sorted(set(lst), key=str))
        grouped['Final Answer'] = grouped['Request Status'].apply(get_final_answer)

        # Merge Final Answer back to main dataframe
        df = df.merge(grouped[['Assigned Request Ids', 'Final Answer']], on='Assigned Request Ids', how='left')
        
        # Check for unmapped requests
        unmapped_count = (df['Final Answer'] == '❌ No matching rule').sum()
        if unmapped_count > 0:
            print(f"   WARNING: {unmapped_count} rows have no matching rule in logic.xlsx")
            print(f"   Unique Request IDs with no rule: {df[df['Final Answer'] == '❌ No matching rule']['Assigned Request Ids'].nunique()}")
            
    except Exception as e:
        print(f"❌ Error computing final status from logic.xlsx: {e}")
        return
    
    # Deduplicate at Request ID + TBM Division + ABM level to get correct counts
    print("🔧 Deduplicating data at Request ID + TBM Division + ABM level...")
    
    # Store original data for validation
    original_request_count = df['Assigned Request Ids'].nunique()
    
    # Deduplicate: Each unique (Request ID + TBM Division + ABM) combination should appear once
    agg_dict = {
        'AFFILIATE': 'first',
        'DIV_NAME': 'first',
        'ABM Name': 'first',
        'ABM EMAIL_ID': 'first',
        'ZBM Terr Code': 'first',
        'ZBM Name': 'first',
        'ZBM EMAIL_ID': 'first',
        'Doctor: Customer Code': 'first',
        'Final Answer': 'first',
        'Rto Reason': 'first',
    }
    
    # Add TBM created by column if it exists and is different
    if tbm_created_by_col and tbm_created_by_col != 'TBM EMAIL_ID':
        agg_dict[tbm_created_by_col] = 'first'
    
    # Add TBM HQ if it exists
    if 'TBM HQ' in df.columns:
        agg_dict['TBM HQ'] = 'first'
    
    # Add ABM HQ if it exists
    if 'ABM HQ' in df.columns:
        agg_dict['ABM HQ'] = 'first'
    
    df_dedup = df.groupby(['Assigned Request Ids', 'TBM Division', 'ABM Terr Code']).agg(agg_dict).reset_index()
    
    print(f"📊 Deduplicated from {len(df)} rows to {len(df_dedup)} unique (Request ID + TBM Division + ABM) combinations")
    print(f"📊 Unique Request IDs after dedup: {df_dedup['Assigned Request Ids'].nunique()}")
    
    # Get unique TBM Divisions
    divisions = df_dedup.groupby('TBM Division').agg({
        'AFFILIATE': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
        'DIV_NAME': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0]
    }).reset_index().sort_values('TBM Division')
    
    print(f"📋 Found {len(divisions)} unique TBM Divisions")
    
    # Debug: Check for any duplicates
    duplicate_codes = divisions['TBM Division'].value_counts()
    if len(duplicate_codes[duplicate_codes > 1]) > 0:
        print(f"WARNING: Found duplicate TBM Division codes after deduplication!")
        print(duplicate_codes[duplicate_codes > 1])
    
    # Debug: Show first few Divisions and their ABMs
    print("\n🔍 Division-ABM Mapping (first 5):")
    for idx, (_, div_row) in enumerate(divisions.head(5).iterrows()):
        div_code = div_row['TBM Division']
        affiliate = div_row['AFFILIATE']
        div_name = div_row['DIV_NAME']
        div_data_temp = df_dedup[df_dedup['TBM Division'] == div_code]
        abms_temp = div_data_temp[['ABM Terr Code', 'ABM Name']].drop_duplicates()
        requests_temp = div_data_temp['Assigned Request Ids'].nunique()
        print(f"   {idx+1}. Division {div_code} ({affiliate} - {div_name}): {len(abms_temp)} ABMs, {requests_temp} requests")
    
    # Create output directory
    timestamp = datetime.now().strftime('%Y%m%d')
    output_dir = f"Division_Reports_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Created output directory: {output_dir}")
    
    # Process each Division
    file_count = 0
    total_validation_errors = 0
    
    for _, div_row in divisions.iterrows():
        div_code = div_row['TBM Division']
        affiliate = div_row['AFFILIATE']
        div_name = div_row['DIV_NAME']
        
        print(f"\n📄 Processing Division: {div_code} - {affiliate} - {div_name}")
        
        # Filter data for this Division (using deduplicated data)
        div_data = df_dedup[df_dedup['TBM Division'] == div_code].copy()
        
        if len(div_data) == 0:
            print(f"No data found for Division: {div_code}")
            continue
        
        # Build hierarchical structure: Division -> ZBM -> ABM -> TBM
        print(f"   📊 Building hierarchical structure (Division -> ZBM -> ABM -> TBM)")
        
        # Get unique ZBMs under this Division
        zbms = div_data.groupby(['ZBM Terr Code', 'ZBM Name']).agg({
            'ZBM EMAIL_ID': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
        }).reset_index().sort_values('ZBM Terr Code')
        
        print(f"   📊 Found {len(zbms)} ZBMs under this Division")
        
        # Create hierarchical summary data
        summary_data = []
        
        # Helper function to calculate metrics for a given dataset
        def calculate_metrics(data_subset, level_name=""):
            """Calculate all metrics for a given data subset"""
            # Calculate unique counts - use nunique() for accurate unique counts
            unique_tbms = data_subset[tbm_created_by_col].nunique()
            unique_hcps = data_subset['Doctor: Customer Code'].nunique()
            unique_requests = data_subset['Assigned Request Ids'].nunique()
            
            # === SECTION A: Request Cancelled Out of Stock ===
            ho_statuses = ['Out of stock', 'On hold', 'Not permitted']
            request_cancelled_out_of_stock = data_subset[data_subset['Final Answer'].isin(ho_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION B: Action Pending at HO ===
            pending_statuses = ['Request Raised', 'Action pending / In Process At HO']
            action_pending_at_ho = data_subset[data_subset['Final Answer'].isin(pending_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION D: Pending for Invoicing ===
            hub_pending_statuses = ['Action pending / In Process At Hub']
            pending_for_invoicing = data_subset[data_subset['Final Answer'].isin(hub_pending_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION E: Pending for Dispatch ===
            dispatch_pending_statuses = ['Dispatch  Pending', 'Dispatch Pending']
            pending_for_dispatch = data_subset[data_subset['Final Answer'].isin(dispatch_pending_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION G: Delivered ===
            delivered_statuses = ['Delivered']
            delivered = data_subset[data_subset['Final Answer'].isin(delivered_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION H: Dispatched & In Transit ===
            transit_statuses = ['Dispatched & In Transit']
            dispatched_in_transit = data_subset[data_subset['Final Answer'].isin(transit_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION I: RTO (Return to Origin) ===
            rto_total = data_subset[data_subset['Final Answer'] == 'Return']['Assigned Request Ids'].nunique()
            
            # RTO Reasons
            unique_request_ids = data_subset[data_subset['Final Answer'] == 'Return']['Assigned Request Ids'].unique()
            
            incomplete_address = 0
            doctor_refused_to_accept = 0
            doctor_non_contactable = 0
            rto_due_to_hold_delivery = 0
            
            for req_id in unique_request_ids:
                req_rows = data_subset[data_subset['Assigned Request Ids'] == req_id]
                rto_col = req_rows['Rto Reason'].astype(str).str.strip().str.lower()
                
                has_incomplete = rto_col.str.contains('incomplete address', na=False, regex=False).any()
                has_refused = rto_col.str.contains('refused to accept', na=False, regex=False).any()
                has_non_contactable = rto_col.str.contains('non contactable', na=False, regex=False).any()
                has_rto_hold_delivery = rto_col.str.contains('hold delivery', na=False, regex=False).any()
                
                if has_incomplete:
                    incomplete_address += 1
                elif has_refused:
                    doctor_refused_to_accept += 1
                elif has_non_contactable:
                    doctor_non_contactable += 1
                elif has_rto_hold_delivery:
                    rto_due_to_hold_delivery += 1
            
            # === CALCULATED FIELDS ===
            requests_dispatched = delivered + dispatched_in_transit + rto_total
            sent_to_hub = pending_for_invoicing + pending_for_dispatch + requests_dispatched
            requests_raised = unique_requests
            
            return {
                'Affiliate': affiliate,
                'Division': div_code,
                'Division Name': div_name,
                '# Unique TBMs': unique_tbms,
                '# Unique HCPs': unique_hcps,
                '# Requests Raised\n(A+B+C)': requests_raised,
                'Request Cancelled / Out of Stock (A)': request_cancelled_out_of_stock,
                'Action pending / In Process At HO (B)': action_pending_at_ho,
                "Sent to HUB ('C)\n(D+E+F)": sent_to_hub,
                'Pending for Invoicing (D)': pending_for_invoicing,
                'Pending for Dispatch (E)': pending_for_dispatch,
                '# Requests Dispatched (F)\n(G+H+I)': requests_dispatched,
                'Delivered (G)': delivered,
                'Dispatched & In Transit (H)': dispatched_in_transit,
                'RTO (I)': rto_total,
                'Incomplete Address': incomplete_address,
                'Doctor Non Contactable': doctor_non_contactable,
                'Doctor Refused to Accept': doctor_refused_to_accept,
                'Hold Delivery': rto_due_to_hold_delivery,
                'Level': level_name,  # For identifying hierarchy level
                'ZBM Code': None,
                'ZBM Name': None,
                'ABM Code': None,
                'ABM Name': None,
                'TBM Code': None,
                'TBM Name': None,
            }
        
        # Process each ZBM
        for _, zbm_row in zbms.iterrows():
            zbm_code = zbm_row['ZBM Terr Code']
            zbm_name = zbm_row['ZBM Name']
            
            # Filter data for this ZBM
            zbm_data = div_data[div_data['ZBM Terr Code'] == zbm_code].copy()
            
            # Calculate ZBM-level metrics
            zbm_metrics = calculate_metrics(zbm_data, 'ZBM')
            zbm_metrics['ZBM Code'] = zbm_code
            zbm_metrics['ZBM Name'] = zbm_name
            summary_data.append(zbm_metrics)
            
            # Get unique ABMs under this ZBM
            abms = zbm_data.groupby(['ABM Terr Code', 'ABM Name']).agg({
                'ABM EMAIL_ID': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
                'TBM HQ': 'first',
                'ABM HQ': 'first' if 'ABM HQ' in zbm_data.columns else lambda x: None
            }).reset_index().sort_values('ABM Terr Code')
            
            # Process each ABM under this ZBM
            for _, abm_row in abms.iterrows():
                abm_code = abm_row['ABM Terr Code']
                abm_name = abm_row['ABM Name']
                
                # Filter data for this ABM
                abm_data = zbm_data[(zbm_data['ABM Terr Code'] == abm_code) & 
                                   (zbm_data['ABM Name'] == abm_name)].copy()
                
                # Calculate ABM-level metrics
                abm_metrics = calculate_metrics(abm_data, 'ABM')
                abm_metrics['ZBM Code'] = zbm_code
                abm_metrics['ZBM Name'] = zbm_name
                abm_metrics['ABM Code'] = abm_code
                abm_metrics['ABM Name'] = abm_name
                summary_data.append(abm_metrics)
                
                # Get unique TBMs under this ABM
                tbms = abm_data.groupby([tbm_created_by_col]).agg({
                    'TBM EMAIL_ID': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0] if 'TBM EMAIL_ID' in abm_data.columns else None,
                    'TBM HQ': 'first' if 'TBM HQ' in abm_data.columns else lambda x: None
                }).reset_index()
                tbms = tbms.rename(columns={tbm_created_by_col: 'TBM_Identifier'})
                
                # Process each TBM under this ABM
                for _, tbm_row in tbms.iterrows():
                    tbm_identifier = tbm_row['TBM_Identifier']
                    
                    # Filter data for this TBM
                    tbm_data = abm_data[abm_data[tbm_created_by_col] == tbm_identifier].copy()
                    
                    # Calculate TBM-level metrics
                    tbm_metrics = calculate_metrics(tbm_data, 'TBM')
                    tbm_metrics['ZBM Code'] = zbm_code
                    tbm_metrics['ZBM Name'] = zbm_name
                    tbm_metrics['ABM Code'] = abm_code
                    tbm_metrics['ABM Name'] = abm_name
                    tbm_metrics['TBM Code'] = tbm_identifier
                    tbm_metrics['TBM Name'] = tbm_identifier  # Use identifier as name if no separate name column
                    summary_data.append(tbm_metrics)
        
        # Calculate Division-level totals (for validation and total row)
        div_metrics = calculate_metrics(div_data, 'Total')
        div_metrics['Level'] = 'Total'
        
        # Validate Division total
        div_total_requests = div_data['Assigned Request Ids'].nunique()
        if div_total_requests != div_metrics['# Requests Raised\n(A+B+C)']:
            print(f"      WARNING: Division {div_code} total mismatch!")
            print(f"      Actual unique requests: {div_total_requests}")
            print(f"      Calculated total: {div_metrics['# Requests Raised\n(A+B+C)']}")
        
        # Add total row
        summary_data.append(div_metrics)
        
        # Create DataFrame for this Division
        div_summary_df = pd.DataFrame(summary_data)
        
        # Validate totals match
        zbm_sum = div_summary_df[div_summary_df['Level'] == 'ZBM']['# Requests Raised\n(A+B+C)'].sum()
        abm_sum = div_summary_df[div_summary_df['Level'] == 'ABM']['# Requests Raised\n(A+B+C)'].sum()
        tbm_sum = div_summary_df[div_summary_df['Level'] == 'TBM']['# Requests Raised\n(A+B+C)'].sum()
        total_sum = div_summary_df[div_summary_df['Level'] == 'Total']['# Requests Raised\n(A+B+C)'].iloc[0]
        
        print(f"   ✅ ZBM sum: {zbm_sum}, ABM sum: {abm_sum}, TBM sum: {tbm_sum}, Total: {total_sum}")
        
        if abs(zbm_sum - total_sum) > 0:
            print(f"      ⚠️  ZBM sum doesn't match total (diff: {abs(zbm_sum - total_sum)})")
            total_validation_errors += 1
        
        # Create Excel file for this Division
        create_division_excel_report(div_code, affiliate, div_name, div_summary_df, output_dir)
        file_count += 1
    
    print(f"\n🎉 Successfully created {file_count} Division reports in directory: {output_dir}")
    print(f"📊 Total Divisions processed: {file_count}")
    if total_validation_errors > 0:
        print(f"WARNING: {total_validation_errors} TBMs had validation errors")
    else:
        print(f"✅ All tallies match perfectly!")

def create_division_excel_report(div_code, affiliate, div_name, summary_df, output_dir):
    """Create Excel report for a specific Division with perfect formatting based on Excel template"""
    
    try:
        # Load Excel template file (not CSV)
        template_file = 'division summary.xlsx'
        
        if not os.path.exists(template_file):
            print(f"   ❌ Template file not found: {template_file}")
            return
        
        # Load the Excel template to preserve formatting
        wb = load_workbook(template_file)
        ws = wb.active

        def get_cell_value_handling_merged(row, col):
            """Get cell value even if it's part of a merged cell"""
            cell = ws.cell(row=row, column=col)
            
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return top_left_cell.value
            
            return cell.value
        
        # Search for header row containing "Affiliate"
        header_row = None
        for row_idx in range(1, 15):
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = get_cell_value_handling_merged(row_idx, col_idx)
                if cell_value and 'Affiliate' in str(cell_value):
                    header_row = row_idx
                    break
            if header_row:
                break
        
        if header_row is None:
            header_row = 3  # Default based on CSV template structure (row 3)
        
        # Find "Total" row
        total_row = None
        for row_idx in range(header_row + 1, min(header_row + 20, ws.max_row + 1)):
            cell_value = get_cell_value_handling_merged(row_idx, 1)
            if cell_value and 'Total' in str(cell_value):
                total_row = row_idx
                break
        
        if total_row is None:
            total_row = header_row + 1  # Default position
            print(f"   ⚠️  'Total' row not found, using default row {total_row}")
        else:
            print(f"   ✅ Found 'Total' row at row {total_row}")
        
        data_start_row = header_row + 1
        
        # Read actual column positions from template header row
        column_mapping = {}
        template_headers = []  # For debugging
        
        # Define mapping rules: (summary_df_column_name, [list of possible header substrings to match])
        # Order matters: more specific matches should come first
        mapping_rules = [
            ('Affiliate', ['Affiliate']),
            ('Division Name', ['Division Name']),  # Check before 'Division' to avoid conflicts
            ('Division', ['Division']),
            ('# Unique TBMs', ['TBMs', '# Unique TBMs', 'Unique TBMs', '# TBMs']),
            ('# Unique HCPs', ['HCPs', '# Unique HCPs', 'Unique HCPs', '# HCPs']),
            ('# Requests Raised\n(A+B+C)', ['Requests Raised', 'Requests raised', '(A+B+C)']),
            ('Request Cancelled / Out of Stock (A)', ['Out of Stock', 'Out of stock', 'Cancelled', '(A)']),
            ('Action pending / In Process At HO (B)', ['Action pending', 'In Process At HO', '(B)']),
            ("Sent to HUB ('C)\n(D+E+F)", ['Sent to HUB', 'Sent to Hub', '(C)', '(D+E+F)']),
            ('Pending for Invoicing (D)', ['Pending for Invoicing', 'Invoicing', '(D)']),
            ('Pending for Dispatch (E)', ['Pending for Dispatch', 'Dispatch', '(E)']),
            ('# Requests Dispatched (F)\n(G+H+I)', ['Requests Dispatched', '(F)', '(G+H+I)']),
            ('Delivered (G)', ['Delivered', '(G)']),
            ('Dispatched & In Transit (H)', ['Dispatched & In Transit', 'Dispatched &amp; In Transit', 'In Transit', '(H)']),
            ('RTO (I)', ['RTO', '(I)']),
            ('Incomplete Address', ['Incomplete Address', 'Incomplete']),
            ('Doctor Non Contactable', ['Doctor Non Contactable', 'Non Contactable', 'Non-contactable']),
            ('Doctor Refused to Accept', ['Refused to Accept', 'refused to accept', 'Refused']),
            ('Hold Delivery', ['Hold Delivery', 'Hold delivery']),
        ]
        
        # First, collect all template headers
        for col_idx in range(1, min(30, ws.max_column + 1)):
            header_val = get_cell_value_handling_merged(header_row, col_idx)
            if header_val:
                header_str = str(header_val).strip()
                template_headers.append((col_idx, header_str))
        
        # Now match each rule to the best column
        for summary_col, search_terms in mapping_rules:
            # Skip if already mapped
            if summary_col in column_mapping:
                continue
            
            # Try to find the best matching column for this rule
            best_match = None
            best_match_score = 0
            
            for col_idx, header_str in template_headers:
                # Skip if this column is already mapped to something else
                if col_idx in column_mapping.values():
                    continue
                
                # Normalize header for matching (remove newlines, extra spaces, case insensitive)
                header_normalized = ' '.join(header_str.replace('\n', ' ').replace('\r', ' ').split()).lower()
                
                # Check if any search term matches
                match_score = 0
                for term in search_terms:
                    term_lower = term.lower()
                    # Exact match gets higher score
                    if term_lower == header_normalized:
                        match_score = 100
                        break
                    elif term_lower in header_normalized:
                        match_score = max(match_score, len(term))
                    elif term_lower in header_str.lower():
                        match_score = max(match_score, len(term) // 2)
                
                # Special handling for Division vs Division Name
                if summary_col == 'Division' and 'name' in header_normalized:
                    match_score = 0  # Don't match "Division Name" to "Division"
                
                if match_score > best_match_score:
                    best_match_score = match_score
                    best_match = col_idx
            
            if best_match and best_match_score > 0:
                column_mapping[summary_col] = best_match
        
        # Debug: Print template headers and mappings
        print(f"   🔍 Template headers found (row {header_row}):")
        for col_idx, header in template_headers[:15]:  # Print first 15
            mapped_to = [k for k, v in column_mapping.items() if v == col_idx]
            mapped_str = f" -> {mapped_to[0]}" if mapped_to else ""
            print(f"      Col {col_idx}: '{header}'{mapped_str}")
        
        print(f"   🔍 Column mappings created: {len(column_mapping)} mappings")
        
        # Check which columns from summary_df have mappings
        missing_mappings = [col for col in summary_df.columns if col not in column_mapping]
        if missing_mappings:
            print(f"   ⚠️  WARNING: {len(missing_mappings)} columns in summary_df have no mapping:")
            for col in missing_mappings:
                print(f"      - '{col}'")
                # Try to find similar headers
                for col_idx, header in template_headers:
                    if any(word.lower() in str(header).lower() for word in col.split() if len(word) > 3):
                        print(f"        (Possible match: Col {col_idx} = '{header}')")
        else:
            print(f"   ✅ All {len(summary_df.columns)} columns have mappings!")

        # Find column for hierarchy identifier (ZBM Name, ABM Name, or Division Name)
        # Priority: ZBM Name > ABM Name > Division Name > first available column
        hierarchy_col = None
        for col_idx, header_str in template_headers:
            header_lower = str(header_str).lower()
            if 'zbm' in header_lower and 'name' in header_lower:
                hierarchy_col = col_idx
                break
        
        if hierarchy_col is None:
            for col_idx, header_str in template_headers:
                header_lower = str(header_str).lower()
                if 'abm' in header_lower and 'name' in header_lower:
                    hierarchy_col = col_idx
                    break
        
        if hierarchy_col is None:
            # Use Division Name column if available
            hierarchy_col = column_mapping.get('Division Name')
        
        if hierarchy_col is None:
            # Use Division column if available
            hierarchy_col = column_mapping.get('Division')
        
        if hierarchy_col is None:
            # Use first column as fallback
            hierarchy_col = 1
            print(f"   ⚠️  No hierarchy column found, using column {hierarchy_col} as fallback")
        
        # Clear existing data rows (between header and total)
        max_data_rows = len(summary_df) + 10
        for r in range(data_start_row, min(data_start_row + max_data_rows, total_row)):
            for c in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                except:
                    pass

        def copy_row_style(src_row_idx, dst_row_idx):
            """Copy formatting from source row to destination row"""
            for c in range(1, ws.max_column + 1):
                try:
                    src = ws.cell(row=src_row_idx, column=c)
                    dst = ws.cell(row=dst_row_idx, column=c)
                    
                    if src.font:
                        dst.font = copy_style(src.font)
                    if src.alignment:
                        dst.alignment = copy_style(src.alignment)
                    if src.border:
                        dst.border = copy_style(src.border)
                    if src.fill:
                        dst.fill = copy_style(src.fill)
                    dst.number_format = src.number_format
                except:
                    pass

        # Separate rows by level
        zbm_rows = summary_df[summary_df['Level'] == 'ZBM'].copy()
        abm_rows = summary_df[summary_df['Level'] == 'ABM'].copy()
        tbm_rows = summary_df[summary_df['Level'] == 'TBM'].copy()
        total_row_data = summary_df[summary_df['Level'] == 'Total'].iloc[0]
        
        # Sort rows to maintain hierarchy
        zbm_rows = zbm_rows.sort_values('ZBM Code')
        abm_rows = abm_rows.sort_values(['ZBM Code', 'ABM Code'])
        tbm_rows = tbm_rows.sort_values(['ZBM Code', 'ABM Code', 'TBM Code'])
        
        # Write hierarchical data rows
        current_row = data_start_row
        template_data_row = data_start_row
        
        # Write ZBM rows
        for idx, (_, zbm_row) in enumerate(zbm_rows.iterrows()):
            target_row = current_row
            copy_row_style(template_data_row, target_row)
            
            # Set hierarchy identifier
            if hierarchy_col:
                zbm_display = f"{zbm_row['ZBM Code']} - {zbm_row['ZBM Name']}"
                ws.cell(row=target_row, column=hierarchy_col, value=zbm_display)
                cell = ws.cell(row=target_row, column=hierarchy_col)
                cell.font = Font(bold=True, name='Arial', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Write data
            for col_name, col_idx in column_mapping.items():
                if col_name in zbm_row.index and col_name not in ['Level', 'ZBM Code', 'ZBM Name', 'ABM Code', 'ABM Name', 'TBM Code', 'TBM Name']:
                    value = zbm_row[col_name]
                    try:
                        cell = ws.cell(row=target_row, column=col_idx)
                        cell.value = value
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0'
                            cell.font = Font(bold=True, name='Arial', size=10)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    except:
                        pass
            
            current_row += 1
            
            # Write ABM rows under this ZBM
            zbm_abms = abm_rows[abm_rows['ZBM Code'] == zbm_row['ZBM Code']]
            for _, abm_row in zbm_abms.iterrows():
                target_row = current_row
                copy_row_style(template_data_row, target_row)
                
                # Set hierarchy identifier (indented)
                if hierarchy_col:
                    abm_display = f"  {abm_row['ABM Code']} - {abm_row['ABM Name']}"
                    ws.cell(row=target_row, column=hierarchy_col, value=abm_display)
                    cell = ws.cell(row=target_row, column=hierarchy_col)
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Write data
                for col_name, col_idx in column_mapping.items():
                    if col_name in abm_row.index and col_name not in ['Level', 'ZBM Code', 'ZBM Name', 'ABM Code', 'ABM Name', 'TBM Code', 'TBM Name']:
                        value = abm_row[col_name]
                        try:
                            cell = ws.cell(row=target_row, column=col_idx)
                            cell.value = value
                            if isinstance(value, (int, float)) and not pd.isna(value):
                                cell.number_format = '0'
                        except:
                            pass
                
                current_row += 1
                
                # Write TBM rows under this ABM
                abm_tbms = tbm_rows[(tbm_rows['ZBM Code'] == abm_row['ZBM Code']) & 
                                   (tbm_rows['ABM Code'] == abm_row['ABM Code'])]
                for _, tbm_row in abm_tbms.iterrows():
                    target_row = current_row
                    copy_row_style(template_data_row, target_row)
                    
                    # Set hierarchy identifier (more indented)
                    if hierarchy_col:
                        tbm_display = f"    {tbm_row['TBM Code']}"
                        ws.cell(row=target_row, column=hierarchy_col, value=tbm_display)
                        cell = ws.cell(row=target_row, column=hierarchy_col)
                        cell.font = Font(name='Arial', size=9)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Write data
                    for col_name, col_idx in column_mapping.items():
                        if col_name in tbm_row.index and col_name not in ['Level', 'ZBM Code', 'ZBM Name', 'ABM Code', 'ABM Name', 'TBM Code', 'TBM Name']:
                            value = tbm_row[col_name]
                            try:
                                cell = ws.cell(row=target_row, column=col_idx)
                                cell.value = value
                                if isinstance(value, (int, float)) and not pd.isna(value):
                                    cell.number_format = '0'
                            except:
                                pass
                    
                    current_row += 1
        
        # Write Total row
        copy_row_style(total_row, total_row)
        
        # Set "Total" text in first column or hierarchy column
        total_label_col = hierarchy_col if hierarchy_col else 1
        ws.cell(row=total_row, column=total_label_col, value="Total")
        cell = ws.cell(row=total_row, column=total_label_col)
        cell.font = Font(bold=True, name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write total data
        values_written = 0
        for col_name, col_idx in column_mapping.items():
            if col_name in total_row_data.index and col_name not in ['Level', 'ZBM Code', 'ZBM Name', 'ABM Code', 'ABM Name', 'TBM Code', 'TBM Name']:
                value = total_row_data[col_name]
                try:
                    cell = ws.cell(row=total_row, column=col_idx)
                    cell.value = value
                    values_written += 1
                    
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        cell.number_format = '0'
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                except Exception as e:
                    print(f"   ⚠️  Warning: Could not set value for column '{col_name}' (col {col_idx}): {e}")
        
        print(f"   ✅ Wrote {len(zbm_rows)} ZBM rows, {len(abm_rows)} ABM rows, {len(tbm_rows)} TBM rows, and 1 Total row")
        print(f"   ✅ Wrote {values_written} values to Total row")

        # Save file
        safe_div_name = str(div_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f"Division_Summary_{div_code}_{safe_div_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb.save(filepath)
        print(f"   ✅ Created: {filename}")
        
    except Exception as e:
        print(f"   ❌ Error creating Excel report for Division {div_code}: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    create_division_hierarchical_reports()
        
   
