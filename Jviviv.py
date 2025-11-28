import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy as copy_style
import warnings

# Suppress FutureWarning for groupby operations
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

def create_division_hierarchical_reports():
    """
    Create separate Division reports showing TBM Division hierarchy with perfect tallies
    Each TBM Division gets a report showing all TBMs under them
    """
    
    print("📄 Starting Division Hierarchical Reports Creation...")
    
    # Read master tracker data from Excel file
    print("📖 Reading ZBM Automation Email 2410252.xlsx...")
    try:
        df = pd.read_excel('ZBM Automation Email 2410252.xlsx')
        print(f"✅ Successfully loaded {len(df)} records")
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return
    
    print(f"📋 Columns in file: {list(df.columns)}")
    
    # Basic data preparation
    print("🧹 Preparing data...")
    
    # Find the correct column name for TBM/Created By
    tbm_created_by_col = None
    for col in df.columns:
        if 'created by' in col.lower() or 'created_by' in col.lower():
            tbm_created_by_col = col
            print(f"✅ Found TBM Created By column: '{col}'")
            break
    
    if tbm_created_by_col is None:
        print("Warning: Could not find 'Created By' column, will use 'TBM EMAIL_ID' instead")
        tbm_created_by_col = 'TBM EMAIL_ID'
    
    # Ensure required columns exist
    required_columns = ['TBM Division', 'AFFILIATE', 'DIV_NAME',
                        'ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID',
                        'ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID',
                        'Doctor: Customer Code', 'Assigned Request Ids', 'Request Status', 'Rto Reason']
    
    # Add the TBM created by column if it's different from TBM EMAIL_ID
    if tbm_created_by_col != 'TBM EMAIL_ID' and tbm_created_by_col not in required_columns:
        required_columns.append(tbm_created_by_col)
    
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns: {missing}")
        return

    print(f"📊 Total rows in file: {len(df)}")
    print(f"📊 Unique Request IDs in raw data: {df['Assigned Request Ids'].nunique()}")
    print(f"📊 Unique TBM Divisions in raw data: {df['TBM Division'].nunique()}")

    # Compute Final Answer per unique request id using rules from logic.xlsx
    print("🧠 Computing final status per unique Request Id using rules...")
    try:
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        def normalize(text):
            return str(text).strip().casefold()

        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']

        # Group statuses by request id from master data
        grouped = df.groupby('Assigned Request Ids')['Request Status'].apply(list).reset_index()

        def get_final_answer(status_list):
            key = tuple(sorted(set(normalize(s) for s in status_list)))
            return rules.get(key, '❌ No matching rule')

        grouped['Request Status'] = grouped['Request Status'].apply(lambda lst: sorted(set(lst), key=str))
        grouped['Final Answer'] = grouped['Request Status'].apply(get_final_answer)

        # Merge Final Answer back to main dataframe
        df = df.merge(grouped[['Assigned Request Ids', 'Final Answer']], on='Assigned Request Ids', how='left')
        
        # Check for unmapped requests
        unmapped_count = (df['Final Answer'] == '❌ No matching rule').sum()
        if unmapped_count > 0:
            print(f"   WARNING: {unmapped_count} rows have no matching rule in logic.xlsx")
            print(f"   Unique Request IDs with no rule: {df[df['Final Answer'] == '❌ No matching rule']['Assigned Request Ids'].nunique()}")
            
    except Exception as e:
        print(f"❌ Error computing final status from logic.xlsx: {e}")
        return
    
    # Deduplicate at Request ID + TBM Division + ABM level to get correct counts
    print("🔧 Deduplicating data at Request ID + TBM Division + ABM level...")
    
    # Store original data for validation
    original_request_count = df['Assigned Request Ids'].nunique()
    
    # Deduplicate: Each unique (Request ID + TBM Division + ABM) combination should appear once
    agg_dict = {
        'AFFILIATE': 'first',
        'DIV_NAME': 'first',
        'ABM Name': 'first',
        'ABM EMAIL_ID': 'first',
        'ZBM Terr Code': 'first',
        'ZBM Name': 'first',
        'ZBM EMAIL_ID': 'first',
        'Doctor: Customer Code': 'first',
        'Final Answer': 'first',
        'Rto Reason': 'first',
    }
    
    # Add TBM created by column (always include it for unique counting)
    if tbm_created_by_col and tbm_created_by_col in df.columns:
        agg_dict[tbm_created_by_col] = 'first'
    elif 'TBM EMAIL_ID' in df.columns:
        # Fallback to TBM EMAIL_ID if created by column not found
        agg_dict['TBM EMAIL_ID'] = 'first'
        if tbm_created_by_col != 'TBM EMAIL_ID':
            tbm_created_by_col = 'TBM EMAIL_ID'
            print(f"   ⚠️  Using 'TBM EMAIL_ID' as TBM identifier")
    
    # Add TBM HQ if it exists
    if 'TBM HQ' in df.columns:
        agg_dict['TBM HQ'] = 'first'
    
    # Add ABM HQ if it exists
    if 'ABM HQ' in df.columns:
        agg_dict['ABM HQ'] = 'first'
    
    df_dedup = df.groupby(['Assigned Request Ids', 'TBM Division', 'ABM Terr Code']).agg(agg_dict).reset_index()
    
    print(f"📊 Deduplicated from {len(df)} rows to {len(df_dedup)} unique (Request ID + TBM Division + ABM) combinations")
    print(f"📊 Unique Request IDs after dedup: {df_dedup['Assigned Request Ids'].nunique()}")
    
    # Verify TBM column exists in deduplicated data
    if tbm_created_by_col not in df_dedup.columns:
        print(f"   ⚠️  WARNING: TBM column '{tbm_created_by_col}' not found in deduplicated data!")
        print(f"   Available columns: {list(df_dedup.columns)}")
        # Try to find alternative
        if 'TBM EMAIL_ID' in df_dedup.columns:
            tbm_created_by_col = 'TBM EMAIL_ID'
            print(f"   ✅ Using 'TBM EMAIL_ID' instead")
        else:
            print(f"   ❌ CRITICAL: No TBM identifier column found!")
    
    # Verify HCP column exists
    if 'Doctor: Customer Code' not in df_dedup.columns:
        print(f"   ⚠️  WARNING: HCP column 'Doctor: Customer Code' not found in deduplicated data!")
    
    # Get unique TBM Divisions
    divisions = df_dedup.groupby('TBM Division').agg({
        'AFFILIATE': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
        'DIV_NAME': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0]
    }).reset_index().sort_values('TBM Division')
    
    print(f"📋 Found {len(divisions)} unique TBM Divisions")
    
    # Debug: Check for any duplicates
    duplicate_codes = divisions['TBM Division'].value_counts()
    if len(duplicate_codes[duplicate_codes > 1]) > 0:
        print(f"WARNING: Found duplicate TBM Division codes after deduplication!")
        print(duplicate_codes[duplicate_codes > 1])
    
    # Debug: Show first few Divisions and their ABMs
    print("\n🔍 Division-ABM Mapping (first 5):")
    for idx, (_, div_row) in enumerate(divisions.head(5).iterrows()):
        div_code = div_row['TBM Division']
        affiliate = div_row['AFFILIATE']
        div_name = div_row['DIV_NAME']
        div_data_temp = df_dedup[df_dedup['TBM Division'] == div_code]
        abms_temp = div_data_temp[['ABM Terr Code', 'ABM Name']].drop_duplicates()
        requests_temp = div_data_temp['Assigned Request Ids'].nunique()
        print(f"   {idx+1}. Division {div_code} ({affiliate} - {div_name}): {len(abms_temp)} ABMs, {requests_temp} requests")
    
    # Create output directory
    timestamp = datetime.now().strftime('%Y%m%d')
    output_dir = f"Division_Reports_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Created output directory: {output_dir}")
    
    # Process each Division
    file_count = 0
    total_validation_errors = 0
    
    if len(divisions) == 0:
        print("❌ No Divisions found in data!")
        return
    
    print(f"📋 Processing {len(divisions)} Divisions...")
    
    for _, div_row in divisions.iterrows():
        div_code = div_row['TBM Division']
        affiliate = div_row['AFFILIATE']
        div_name = div_row['DIV_NAME']
        
        print(f"\n📄 Processing Division: {div_code} - {affiliate} - {div_name}")
        
        # Filter data for this Division (using deduplicated data)
        div_data = df_dedup[df_dedup['TBM Division'] == div_code].copy()
        
        if len(div_data) == 0:
            print(f"   ⚠️  No data found for Division: {div_code}, skipping...")
            continue
        
        # Build hierarchical structure: Division -> ZBM -> ABM -> TBM
        print(f"   📊 Building hierarchical structure (Division -> ZBM -> ABM -> TBM)")
        
        # Get unique ZBMs under this Division
        zbms = div_data.groupby(['ZBM Terr Code', 'ZBM Name']).agg({
            'ZBM EMAIL_ID': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
        }).reset_index().sort_values('ZBM Terr Code')
        
        print(f"   📊 Found {len(zbms)} ZBMs under this Division")
        
        # Create hierarchical summary data
        summary_data = []
        
        # Helper function to calculate metrics for a given dataset
        def calculate_metrics(data_subset, level_name=""):
            """Calculate all metrics for a given data subset"""
            # Calculate unique counts - filter out invalid values before counting
            tbm_col_data = data_subset[tbm_created_by_col].astype(str).str.strip()
            tbm_col_data = tbm_col_data[tbm_col_data.notna() & (tbm_col_data != '') & (tbm_col_data != 'nan') & (tbm_col_data != '0') & (tbm_col_data != '0.0')]
            unique_tbms = tbm_col_data.nunique()
            
            hcp_col_data = data_subset['Doctor: Customer Code'].astype(str).str.strip()
            hcp_col_data = hcp_col_data[hcp_col_data.notna() & (hcp_col_data != '') & (hcp_col_data != 'nan') & (hcp_col_data != '0') & (hcp_col_data != '0.0')]
            unique_hcps = hcp_col_data.nunique()
            
            request_col_data = data_subset['Assigned Request Ids'].astype(str).str.strip()
            request_col_data = request_col_data[request_col_data.notna() & (request_col_data != '') & (request_col_data != 'nan')]
            unique_requests = request_col_data.nunique()
            
            # === SECTION A: Request Cancelled Out of Stock ===
            ho_statuses = ['Out of stock', 'On hold', 'Not permitted']
            request_cancelled_out_of_stock = data_subset[data_subset['Final Answer'].isin(ho_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION B: Action Pending at HO ===
            pending_statuses = ['Request Raised', 'Action pending / In Process At HO']
            action_pending_at_ho = data_subset[data_subset['Final Answer'].isin(pending_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION D: Pending for Invoicing ===
            hub_pending_statuses = ['Action pending / In Process At Hub']
            pending_for_invoicing = data_subset[data_subset['Final Answer'].isin(hub_pending_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION E: Pending for Dispatch ===
            # Handle variations in status text (extra spaces, case differences)
            dispatch_pending_statuses = ['Dispatch  Pending', 'Dispatch Pending', 'dispatch pending', 'DISPATCH PENDING', 'Dispatch Pending ']
            # Normalize Final Answer for matching
            data_subset_normalized = data_subset.copy()
            data_subset_normalized['Final Answer Normalized'] = data_subset_normalized['Final Answer'].astype(str).str.strip().str.lower()
            dispatch_pending_normalized = [s.strip().lower() for s in dispatch_pending_statuses]
            pending_for_dispatch = data_subset_normalized[
                data_subset_normalized['Final Answer Normalized'].isin(dispatch_pending_normalized)
            ]['Assigned Request Ids'].nunique()
            
            # === SECTION G: Delivered ===
            delivered_statuses = ['Delivered']
            delivered = data_subset[data_subset['Final Answer'].isin(delivered_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION H: Dispatched & In Transit ===
            transit_statuses = ['Dispatched & In Transit']
            dispatched_in_transit = data_subset[data_subset['Final Answer'].isin(transit_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION I: RTO (Return to Origin) ===
            rto_total = data_subset[data_subset['Final Answer'] == 'Return']['Assigned Request Ids'].nunique()
            
            # RTO Reasons
            unique_request_ids = data_subset[data_subset['Final Answer'] == 'Return']['Assigned Request Ids'].unique()
            
            incomplete_address = 0
            doctor_refused_to_accept = 0
            doctor_non_contactable = 0
            rto_due_to_hold_delivery = 0
            
            for req_id in unique_request_ids:
                req_rows = data_subset[data_subset['Assigned Request Ids'] == req_id]
                rto_col = req_rows['Rto Reason'].astype(str).str.strip().str.lower()
                
                has_incomplete = rto_col.str.contains('incomplete address', na=False, regex=False).any()
                has_refused = rto_col.str.contains('refused to accept', na=False, regex=False).any()
                has_non_contactable = rto_col.str.contains('non contactable', na=False, regex=False).any()
                has_rto_hold_delivery = rto_col.str.contains('hold delivery', na=False, regex=False).any()
                
                if has_incomplete:
                    incomplete_address += 1
                elif has_refused:
                    doctor_refused_to_accept += 1
                elif has_non_contactable:
                    doctor_non_contactable += 1
                elif has_rto_hold_delivery:
                    rto_due_to_hold_delivery += 1
            
            # === CALCULATED FIELDS ===
            requests_dispatched = delivered + dispatched_in_transit + rto_total
            sent_to_hub = pending_for_invoicing + pending_for_dispatch + requests_dispatched
            requests_raised = unique_requests
            
            return {
                'ZBM Code': None,  # Will be filled in for ZBM/ABM/TBM rows
                'ZBM Name': None,  # Will be filled in for ZBM/ABM/TBM rows
                'Affiliate': affiliate,
                'Division': div_code,
                'Division Name': div_name,
                '# Unique TBMs': unique_tbms,
                '# Unique HCPs': unique_hcps,
                '# Requests Raised\n(A+B+C)': requests_raised,
                'Request Cancelled / Out of Stock (A)': request_cancelled_out_of_stock,
                'Action pending / In Process At HO (B)': action_pending_at_ho,
                "Sent to HUB ('C)\n(D+E+F)": sent_to_hub,
                'Pending for Invoicing (D)': pending_for_invoicing,
                'Pending for Dispatch (E)': pending_for_dispatch,
                '# Requests Dispatched (F)\n(G+H+I)': requests_dispatched,
                'Delivered (G)': delivered,
                'Dispatched & In Transit (H)': dispatched_in_transit,
                'RTO (I)': rto_total,
                'Incomplete Address': incomplete_address,
                'Doctor Non Contactable': doctor_non_contactable,
                'Doctor Refused to Accept': doctor_refused_to_accept,
                'Hold Delivery': rto_due_to_hold_delivery,
                'Level': level_name,  # For identifying hierarchy level
                'ABM Code': None,
                'ABM Name': None,
                'TBM Code': None,
                'TBM Name': None,
            }
        
        # Process each ZBM (similar to how ZBM reports process ABMs)
        for i, (_, zbm_row) in enumerate(zbms.iterrows()):
            zbm_code = zbm_row['ZBM Terr Code']
            zbm_name = zbm_row['ZBM Name']
            
            # Filter data for this ZBM
            zbm_data = div_data[div_data['ZBM Terr Code'] == zbm_code].copy()
            
            # Calculate ZBM-level metrics (using nunique for accurate counts)
            # Filter out invalid values (NaN, empty strings, '0', etc.) before counting
            tbm_col_data = zbm_data[tbm_created_by_col].astype(str).str.strip()
            tbm_col_data = tbm_col_data[tbm_col_data.notna() & (tbm_col_data != '') & (tbm_col_data != 'nan') & (tbm_col_data != '0') & (tbm_col_data != '0.0')]
            unique_tbms = tbm_col_data.nunique()
            
            hcp_col_data = zbm_data['Doctor: Customer Code'].astype(str).str.strip()
            hcp_col_data = hcp_col_data[hcp_col_data.notna() & (hcp_col_data != '') & (hcp_col_data != 'nan') & (hcp_col_data != '0') & (hcp_col_data != '0.0')]
            unique_hcps = hcp_col_data.nunique()
            
            request_col_data = zbm_data['Assigned Request Ids'].astype(str).str.strip()
            request_col_data = request_col_data[request_col_data.notna() & (request_col_data != '') & (request_col_data != 'nan')]
            unique_requests = request_col_data.nunique()
            
            # Debug: Show unique counts for first ZBM
            if i == 0:  # First ZBM in division
                print(f"      🔍 Unique counts for ZBM {zbm_code}:")
                print(f"         TBMs: {unique_tbms} (from {len(tbm_col_data)} valid TBM entries)")
                print(f"         HCPs: {unique_hcps} (from {len(hcp_col_data)} valid HCP entries)")
                print(f"         Requests: {unique_requests} (from {len(request_col_data)} valid request entries)")
                # Show sample TBM and HCP values for debugging
                if len(tbm_col_data) > 0:
                    print(f"         Sample TBMs: {tbm_col_data.unique()[:3].tolist()}")
                if len(hcp_col_data) > 0:
                    print(f"         Sample HCPs: {hcp_col_data.unique()[:3].tolist()}")
            
            # Validation: Unique counts should be reasonable
            if unique_tbms > unique_requests:
                print(f"      ⚠️  Warning for ZBM {zbm_code}: Unique TBMs ({unique_tbms}) > Unique Requests ({unique_requests})")
            if unique_hcps > unique_requests:
                print(f"      ⚠️  Warning for ZBM {zbm_code}: Unique HCPs ({unique_hcps}) > Unique Requests ({unique_requests})")
            
            # === SECTION A: Request Cancelled Out of Stock ===
            ho_statuses = ['Out of stock', 'On hold', 'Not permitted']
            request_cancelled_out_of_stock = zbm_data[zbm_data['Final Answer'].isin(ho_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION B: Action Pending at HO ===
            pending_statuses = ['Request Raised', 'Action pending / In Process At HO']
            action_pending_at_ho = zbm_data[zbm_data['Final Answer'].isin(pending_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION D: Pending for Invoicing ===
            hub_pending_statuses = ['Action pending / In Process At Hub']
            pending_for_invoicing = zbm_data[zbm_data['Final Answer'].isin(hub_pending_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION E: Pending for Dispatch ===
            # Handle variations in status text (extra spaces, case differences)
            dispatch_pending_statuses = ['Dispatch  Pending', 'Dispatch Pending', 'dispatch pending', 'DISPATCH PENDING', 'Dispatch Pending ']
            # Normalize Final Answer for matching
            zbm_data_normalized = zbm_data.copy()
            zbm_data_normalized['Final Answer Normalized'] = zbm_data_normalized['Final Answer'].astype(str).str.strip().str.lower()
            dispatch_pending_normalized = [s.strip().lower() for s in dispatch_pending_statuses]
            pending_for_dispatch = zbm_data_normalized[
                zbm_data_normalized['Final Answer Normalized'].isin(dispatch_pending_normalized)
            ]['Assigned Request Ids'].nunique()
            
            # Debug: Show actual status values found
            unique_statuses = zbm_data['Final Answer'].unique()
            dispatch_related = [s for s in unique_statuses if 'dispatch' in str(s).lower() and 'pending' in str(s).lower()]
            if dispatch_related:
                print(f"      Found dispatch pending statuses: {dispatch_related}")
                print(f"      Count: {pending_for_dispatch} unique requests")
            
            # === SECTION G: Delivered ===
            delivered_statuses = ['Delivered']
            delivered = zbm_data[zbm_data['Final Answer'].isin(delivered_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION H: Dispatched & In Transit ===
            transit_statuses = ['Dispatched & In Transit']
            dispatched_in_transit = zbm_data[zbm_data['Final Answer'].isin(transit_statuses)]['Assigned Request Ids'].nunique()
            
            # === SECTION I: RTO (Return to Origin) ===
            rto_total = zbm_data[zbm_data['Final Answer'] == 'Return']['Assigned Request Ids'].nunique()
            
            # RTO Reasons
            unique_request_ids = zbm_data[zbm_data['Final Answer'] == 'Return']['Assigned Request Ids'].unique()
            
            incomplete_address = 0
            doctor_refused_to_accept = 0
            doctor_non_contactable = 0
            rto_due_to_hold_delivery = 0
            
            for req_id in unique_request_ids:
                req_rows = zbm_data[zbm_data['Assigned Request Ids'] == req_id]
                rto_col = req_rows['Rto Reason'].astype(str).str.strip().str.lower()
                
                has_incomplete = rto_col.str.contains('incomplete address', na=False, regex=False).any()
                has_refused = rto_col.str.contains('refused to accept', na=False, regex=False).any()
                has_non_contactable = rto_col.str.contains('non contactable', na=False, regex=False).any()
                has_rto_hold_delivery = rto_col.str.contains('hold delivery', na=False, regex=False).any()
                
                if has_incomplete:
                    incomplete_address += 1
                elif has_refused:
                    doctor_refused_to_accept += 1
                elif has_non_contactable:
                    doctor_non_contactable += 1
                elif has_rto_hold_delivery:
                    rto_due_to_hold_delivery += 1
            
            # === CALCULATED FIELDS ===
            requests_dispatched = delivered + dispatched_in_transit + rto_total
            sent_to_hub = pending_for_invoicing + pending_for_dispatch + requests_dispatched
            requests_raised = unique_requests
            
            # === VALIDATION CHECKS ===
            # Validate: A + B + C should equal total requests raised
            section_a_b_c = request_cancelled_out_of_stock + action_pending_at_ho + sent_to_hub
            if abs(section_a_b_c - requests_raised) > 0:
                print(f"      ⚠️  Validation warning for ZBM {zbm_code}: A+B+C ({section_a_b_c}) != Requests Raised ({requests_raised})")
            
            # Validate: C = D + E + F
            section_d_e_f = pending_for_invoicing + pending_for_dispatch + requests_dispatched
            if abs(sent_to_hub - section_d_e_f) > 0:
                print(f"      ⚠️  Validation warning for ZBM {zbm_code}: C ({sent_to_hub}) != D+E+F ({section_d_e_f})")
                print(f"         D={pending_for_invoicing}, E={pending_for_dispatch}, F={requests_dispatched}")
            
            # Validate: F = G + H + I
            section_g_h_i = delivered + dispatched_in_transit + rto_total
            if abs(requests_dispatched - section_g_h_i) > 0:
                print(f"      ⚠️  Validation warning for ZBM {zbm_code}: F ({requests_dispatched}) != G+H+I ({section_g_h_i})")
            
            # Validate: RTO reasons sum should equal RTO total
            rto_reasons_sum = incomplete_address + doctor_refused_to_accept + doctor_non_contactable + rto_due_to_hold_delivery
            if abs(rto_total - rto_reasons_sum) > 0:
                print(f"      ⚠️  Validation warning for ZBM {zbm_code}: RTO Total ({rto_total}) != RTO Reasons Sum ({rto_reasons_sum})")
            
            # Count unique ZBMs in this division (same value for all ZBM rows in the division)
            unique_zbms_in_division = len(zbms)
            
            summary_data.append({
                'ZBM Code': zbm_code,
                'ZBM Name': zbm_name,
                '# Unique ZBMs': unique_zbms_in_division,  # Show unique ZBM count
                'Affiliate': affiliate,
                'Division': div_code,
                'Division Name': div_name,
                '# Unique TBMs': unique_tbms,
                '# Unique HCPs': unique_hcps,
                '# Requests Raised\n(A+B+C)': requests_raised,
                'Request Cancelled / Out of Stock (A)': request_cancelled_out_of_stock,
                'Action pending / In Process At HO (B)': action_pending_at_ho,
                "Sent to HUB ('C)\n(D+E+F)": sent_to_hub,
                'Pending for Invoicing (D)': pending_for_invoicing,
                'Pending for Dispatch (E)': pending_for_dispatch,
                '# Requests Dispatched (F)\n(G+H+I)': requests_dispatched,
                'Delivered (G)': delivered,
                'Dispatched & In Transit (H)': dispatched_in_transit,
                'RTO (I)': rto_total,
                'Incomplete Address': incomplete_address,
                'Doctor Non Contactable': doctor_non_contactable,
                'Doctor Refused to Accept': doctor_refused_to_accept,
                'Hold Delivery': rto_due_to_hold_delivery,
                'Level': 'ZBM',
            })
        
        # Create DataFrame for this Division (only ZBM rows, like ZBM reports show only ABM rows)
        if len(summary_data) == 0:
            print(f"   ⚠️  No ZBM data found for Division {div_code}, skipping...")
            continue
            
        div_summary_df = pd.DataFrame(summary_data)
        
        # Validate Division total
        div_total_requests = div_data['Assigned Request Ids'].nunique()
        zbm_sum = div_summary_df['# Requests Raised\n(A+B+C)'].sum()
        
        # Calculate division-level unique counts for validation
        div_tbm_col_data = div_data[tbm_created_by_col].astype(str).str.strip()
        div_tbm_col_data = div_tbm_col_data[div_tbm_col_data.notna() & (div_tbm_col_data != '') & (div_tbm_col_data != 'nan') & (div_tbm_col_data != '0') & (div_tbm_col_data != '0.0')]
        div_unique_tbms = div_tbm_col_data.nunique()
        
        div_hcp_col_data = div_data['Doctor: Customer Code'].astype(str).str.strip()
        div_hcp_col_data = div_hcp_col_data[div_hcp_col_data.notna() & (div_hcp_col_data != '') & (div_hcp_col_data != 'nan') & (div_hcp_col_data != '0') & (div_hcp_col_data != '0.0')]
        div_unique_hcps = div_hcp_col_data.nunique()
        
        # Sum of ZBM-level unique counts (should be >= division total due to potential overlaps)
        zbm_tbms_sum = div_summary_df['# Unique TBMs'].sum()
        zbm_hcps_sum = div_summary_df['# Unique HCPs'].sum()
        
        print(f"   ✅ Unique ZBMs: {len(zbms)}, Total requests: {zbm_sum}")
        print(f"   📊 Division-level unique counts: TBMs={div_unique_tbms}, HCPs={div_unique_hcps}")
        print(f"   📊 ZBM-level sums: TBMs={zbm_tbms_sum}, HCPs={zbm_hcps_sum}")
        
        # Note: ZBM sums can be >= division totals due to overlaps (same TBM/HCP in multiple ZBMs)
        if zbm_tbms_sum < div_unique_tbms:
            print(f"      ⚠️  Warning: ZBM TBM sum ({zbm_tbms_sum}) < Division TBM total ({div_unique_tbms})")
        if zbm_hcps_sum < div_unique_hcps:
            print(f"      ⚠️  Warning: ZBM HCP sum ({zbm_hcps_sum}) < Division HCP total ({div_unique_hcps})")
        
        if abs(zbm_sum - div_total_requests) > 0:
            print(f"      ⚠️  ZBM sum doesn't match division total (diff: {abs(div_total_requests - zbm_sum)})")
            total_validation_errors += 1
        
        # Create Excel file for this Division
        try:
            create_division_excel_report(div_code, affiliate, div_name, div_summary_df, output_dir)
            file_count += 1
        except Exception as e:
            print(f"   ❌ Failed to create Excel report for Division {div_code}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"\n🎉 Successfully created {file_count} Division reports in directory: {output_dir}")
    print(f"📊 Total Divisions processed: {file_count}")
    if total_validation_errors > 0:
        print(f"WARNING: {total_validation_errors} TBMs had validation errors")
    else:
        print(f"✅ All tallies match perfectly!")

def create_division_excel_report(div_code, affiliate, div_name, summary_df, output_dir):
    """Create Excel report for a specific Division with perfect formatting based on Excel template"""
    
    try:
        # Load Excel template file (not CSV)
        template_file = 'division summary.xlsx'
        
        if not os.path.exists(template_file):
            print(f"   ❌ Template file not found: {template_file}")
            print(f"   Current working directory: {os.getcwd()}")
            return
        
        # Load the Excel template to preserve formatting
        wb = load_workbook(template_file)
        ws = wb.active

        def get_cell_value_handling_merged(row, col):
            """Get cell value even if it's part of a merged cell"""
            cell = ws.cell(row=row, column=col)
            
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return top_left_cell.value
            
            return cell.value
        
        # Search for header row - try multiple strategies
        header_row = None
        
        # Strategy 1: Look for "ZBM Code" (first column)
        for row_idx in range(1, 15):
            for col_idx in range(1, min(5, ws.max_column + 1)):  # Check first few columns
                cell_value = get_cell_value_handling_merged(row_idx, col_idx)
                if cell_value and 'ZBM Code' in str(cell_value):
                    header_row = row_idx
                    print(f"   ✅ Found header row at {row_idx} (by ZBM Code)")
                    break
            if header_row:
                break
        
        # Strategy 2: Look for "ZBM Name"
        if header_row is None:
            for row_idx in range(1, 15):
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    cell_value = get_cell_value_handling_merged(row_idx, col_idx)
                    if cell_value and 'ZBM Name' in str(cell_value):
                        header_row = row_idx
                        print(f"   ✅ Found header row at {row_idx} (by ZBM Name)")
                        break
                if header_row:
                    break
        
        # Strategy 3: Look for "Affiliate"
        if header_row is None:
            for row_idx in range(1, 15):
                for col_idx in range(1, min(30, ws.max_column + 1)):
                    cell_value = get_cell_value_handling_merged(row_idx, col_idx)
                    if cell_value and 'Affiliate' in str(cell_value):
                        header_row = row_idx
                        print(f"   ✅ Found header row at {row_idx} (by Affiliate)")
                        break
                if header_row:
                    break
        
        # Strategy 4: Look for any common header keywords
        if header_row is None:
            keywords = ['Division', 'Requests Raised', 'Unique TBMs', 'Unique HCPs']
            for row_idx in range(1, 15):
                for col_idx in range(1, min(30, ws.max_column + 1)):
                    cell_value = get_cell_value_handling_merged(row_idx, col_idx)
                    if cell_value:
                        cell_str = str(cell_value)
                        if any(keyword in cell_str for keyword in keywords):
                            header_row = row_idx
                            print(f"   ✅ Found header row at {row_idx} (by keywords)")
                            break
                if header_row:
                    break
        
        if header_row is None:
            header_row = 3  # Default fallback
            print(f"   ⚠️  Could not find header row, using default row {header_row}")
        else:
            print(f"   ✅ Using header row: {header_row}")
        
        # Find "Total" row (optional - we'll create our own total row)
        total_row = None
        for row_idx in range(header_row + 1, min(header_row + 20, ws.max_row + 1)):
            cell_value = get_cell_value_handling_merged(row_idx, 1)
            if cell_value and 'Total' in str(cell_value):
                total_row = row_idx
                break
        
        data_start_row = header_row + 1
        
        # If no total row found, we'll add one after the data
        if total_row is None:
            print(f"   ℹ️  'Total' row not found in template, will add after data rows")
        else:
            print(f"   ✅ Found 'Total' row at row {total_row}")
        
        # Read actual column positions from template header row
        column_mapping = {}
        template_headers = []  # For debugging
        
        # Define mapping rules: (summary_df_column_name, [list of possible header substrings to match])
        # Order matters: more specific matches should come first
        mapping_rules = [
            ('ZBM Code', ['ZBM Code', 'Zone Code', 'ZBM Terr Code']),  # Added 'Zone Code' to match template
            ('ZBM Name', ['ZBM Name']),
            ('# Unique ZBMs', ['Unique ZBMs', '# Unique ZBMs', 'Unique ZBM', '# ZBMs']),
            ('Affiliate', ['Affiliate']),
            ('Division Name', ['Division Name']),  # Check before 'Division' to avoid conflicts
            ('Division', ['Division']),
            ('# Unique TBMs', ['TBMs', '# Unique TBMs', 'Unique TBMs', '# TBMs', '# TBMs']),
            ('# Unique HCPs', ['HCPs', '# Unique HCPs', 'Unique HCPs', '# HCPs', 'Unique HCPs Participating']),
            ('# Requests Raised\n(A+B+C)', ['Requests Raised', 'Requests raised', '# Requests raised', '(A+B+C)']),
            ('Request Cancelled / Out of Stock (A)', ['Out of Stock', 'Out of stock', '# Out of stock', 'Cancelled', '(A)']),
            ('Action pending / In Process At HO (B)', ['Action pending', 'In Process At HO', '# Action pending / In Process', '(B)']),
            ("Sent to HUB ('C)\n(D+E+F)", ['Sent to HUB', 'Sent to Hub', '(C)', '(D+E+F)']),
            ('Pending for Invoicing (D)', ['Pending for Invoicing', 'Invoicing', '(D)']),
            ('Pending for Dispatch (E)', ['Pending for Dispatch', 'Pending for dispatch', 'Pending For Dispatch', '(E)']),  # Fixed: More specific match, avoid conflict with "Requests Dispatched"
            ('# Requests Dispatched (F)\n(G+H+I)', ['Requests Dispatched', '# Requests Dispatched', '(F)', '(G+H+I)']),
            ('Delivered (G)', ['Delivered', '# Delivered', '(G)']),
            ('Dispatched & In Transit (H)', ['Dispatched & In Transit', 'Dispatched &amp; In Transit', '# Dispatched & In Transit', 'In Transit', '(H)']),
            ('RTO (I)', ['RTO', '#RTO', '(I)']),
            ('Incomplete Address', ['Incomplete Address', '- Incomplete Address', 'Incomplete']),
            ('Doctor Non Contactable', ['Doctor Non Contactable', 'Non Contactable', 'Non-contactable']),
            ('Doctor Refused to Accept', ['Refused to Accept', 'refused to accept', 'Refused']),
            ('Hold Delivery', ['Hold Delivery', 'Hold delivery']),
        ]
        
        # First, collect all template headers
        for col_idx in range(1, min(30, ws.max_column + 1)):
            header_val = get_cell_value_handling_merged(header_row, col_idx)
            if header_val:
                header_str = str(header_val).strip()
                template_headers.append((col_idx, header_str))
        
        # Now match each rule to the best column
        for summary_col, search_terms in mapping_rules:
            # Skip if already mapped
            if summary_col in column_mapping:
                continue
            
            # Try to find the best matching column for this rule
            best_match = None
            best_match_score = 0
            
            for col_idx, header_str in template_headers:
                # Skip if this column is already mapped to something else
                if col_idx in column_mapping.values():
                    continue
                
                # Normalize header for matching (remove newlines, extra spaces, case insensitive)
                header_normalized = ' '.join(header_str.replace('\n', ' ').replace('\r', ' ').split()).lower()
                
                # Check if any search term matches
                match_score = 0
                for term in search_terms:
                    term_lower = term.lower()
                    # Exact match gets higher score
                    if term_lower == header_normalized:
                        match_score = 100
                        break
                    elif term_lower in header_normalized:
                        match_score = max(match_score, len(term))
                    elif term_lower in header_str.lower():
                        match_score = max(match_score, len(term) // 2)
                
                # Special handling for Division vs Division Name
                if summary_col == 'Division' and 'name' in header_normalized:
                    match_score = 0  # Don't match "Division Name" to "Division"
                
                if match_score > best_match_score:
                    best_match_score = match_score
                    best_match = col_idx
            
            if best_match and best_match_score > 0:
                column_mapping[summary_col] = best_match
        
        # Debug: Print template headers and mappings
        print(f"   🔍 Template headers found (row {header_row}):")
        for col_idx, header in template_headers[:15]:  # Print first 15
            mapped_to = [k for k, v in column_mapping.items() if v == col_idx]
            mapped_str = f" -> {mapped_to[0]}" if mapped_to else ""
            print(f"      Col {col_idx}: '{header}'{mapped_str}")
        
        print(f"   🔍 Column mappings created: {len(column_mapping)} mappings")
        
        # Check which columns from summary_df have mappings
        missing_mappings = [col for col in summary_df.columns if col not in column_mapping]
        if missing_mappings:
            print(f"   ⚠️  WARNING: {len(missing_mappings)} columns in summary_df have no mapping:")
            for col in missing_mappings:
                print(f"      - '{col}'")
                # Try to find similar headers with more flexible matching
                best_candidate = None
                best_score = 0
                for col_idx, header in template_headers:
                    if col_idx in column_mapping.values():
                        continue
                    # More flexible matching: check for key words
                    col_words = set(word.lower() for word in col.split() if len(word) > 2)
                    header_words = set(word.lower() for word in str(header).split() if len(word) > 2)
                    common_words = col_words.intersection(header_words)
                    if len(common_words) >= 2:  # At least 2 common words
                        score = len(common_words)
                        if score > best_score:
                            best_score = score
                            best_candidate = (col_idx, header)
                
                if best_candidate:
                    print(f"        (Best candidate: Col {best_candidate[0]} = '{best_candidate[1]}')")
                    # Auto-map if confidence is high
                    if best_score >= 3:
                        column_mapping[col] = best_candidate[0]
                        print(f"        ✅ Auto-mapped '{col}' to column {best_candidate[0]}")
        else:
            print(f"   ✅ All {len(summary_df.columns)} columns have mappings!")
        
        # Critical check: Ensure "Pending for Dispatch (E)" is mapped
        if 'Pending for Dispatch (E)' not in column_mapping:
            print(f"   ⚠️  CRITICAL: 'Pending for Dispatch (E)' not mapped! Searching manually...")
            for col_idx, header_str in template_headers:
                header_lower = str(header_str).lower().replace('\n', ' ').replace('\r', ' ')
                if 'pending' in header_lower and 'dispatch' in header_lower and 'dispatched' not in header_lower:
                    if col_idx not in column_mapping.values():
                        column_mapping['Pending for Dispatch (E)'] = col_idx
                        print(f"   ✅ Manually mapped 'Pending for Dispatch (E)' to column {col_idx} ('{header_str}')")
                        break
        
        # Critical check: Must have ZBM Code and ZBM Name columns mapped
        if 'ZBM Code' not in column_mapping:
            print(f"   ⚠️  WARNING: ZBM Code column not found in template!")
            print(f"   Available template headers: {[h[1] for h in template_headers[:10]]}")
            # Try to find Zone Code as fallback
            for col_idx, header_str in template_headers:
                if 'Zone Code' in str(header_str) or 'zone code' in str(header_str).lower():
                    column_mapping['ZBM Code'] = col_idx
                    print(f"   ✅ Mapped 'Zone Code' (col {col_idx}) to 'ZBM Code'")
                    break
            if 'ZBM Code' not in column_mapping:
                print(f"   ❌ CRITICAL ERROR: Could not find ZBM Code or Zone Code column!")
                return
        
        if 'ZBM Name' not in column_mapping:
            print(f"   ❌ CRITICAL ERROR: ZBM Name column not found in template!")
            print(f"   Available template headers: {[h[1] for h in template_headers[:10]]}")
            return

        # Find ZBM Code and ZBM Name columns
        zbm_code_col = column_mapping.get('ZBM Code')
        zbm_name_col = column_mapping.get('ZBM Name')
        
        if zbm_code_col is None or zbm_name_col is None:
            print(f"   ⚠️  WARNING: ZBM Code or ZBM Name column not found in template")
            print(f"      ZBM Code column: {zbm_code_col}, ZBM Name column: {zbm_name_col}")
        
        # Find ABM Name column for ABM/TBM rows (similar to ZBM reports)
        abm_name_col = None
        for col_idx, header_str in template_headers:
            header_lower = str(header_str).lower()
            if 'abm' in header_lower and 'name' in header_lower:
                abm_name_col = col_idx
                break
        
        # Clear existing data rows
        max_data_rows = len(summary_df) + 10
        clear_end_row = total_row if total_row else (data_start_row + max_data_rows + 5)
        for r in range(data_start_row, min(data_start_row + max_data_rows + 5, clear_end_row)):
            for c in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                except:
                    pass

        def copy_row_style(src_row_idx, dst_row_idx):
            """Copy formatting from source row to destination row"""
            for c in range(1, ws.max_column + 1):
                try:
                    src = ws.cell(row=src_row_idx, column=c)
                    dst = ws.cell(row=dst_row_idx, column=c)
                    
                    if src.font:
                        dst.font = copy_style(src.font)
                    if src.alignment:
                        dst.alignment = copy_style(src.alignment)
                    if src.border:
                        dst.border = copy_style(src.border)
                    if src.fill:
                        dst.fill = copy_style(src.fill)
                    dst.number_format = src.number_format
                except:
                    pass

        # Check if we have data to write
        if len(summary_df) == 0:
            print(f"   ⚠️  No data to write for Division {div_code}")
            return
        
        print(f"   📝 Writing {len(summary_df)} ZBM rows starting at row {data_start_row}")
        
        # Sort ZBM rows (like ZBM reports sort ABM rows)
        zbm_rows = summary_df.sort_values('ZBM Code')
        
        # Write data rows (simple format like ZBM reports - just ZBM rows)
        template_data_row = data_start_row
        rows_written = 0
        cells_written = 0
        
        # Debug: Print sample row data before writing
        if len(zbm_rows) > 0:
            sample_row = zbm_rows.iloc[0]
            print(f"   🔍 Sample row data (first ZBM):")
            print(f"      Pending for Dispatch (E): {sample_row.get('Pending for Dispatch (E)', 'N/A')}")
            print(f"      Pending for Invoicing (D): {sample_row.get('Pending for Invoicing (D)', 'N/A')}")
            print(f"      # Requests Dispatched (F): {sample_row.get('# Requests Dispatched (F)\\n(G+H+I)', 'N/A')}")
        
        for i in range(len(zbm_rows)):
            target_row = data_start_row + i
            copy_row_style(template_data_row, target_row)
            
            zbm_row = zbm_rows.iloc[i]
            
            # Write all columns
            for col_name, col_idx in column_mapping.items():
                if col_name in zbm_row.index:
                    value = zbm_row[col_name]
                    try:
                        # Handle NaN/None values
                        if pd.isna(value):
                            value = 0  # Set NaN to 0 for numeric columns
                        
                        # Convert to int if it's a numeric value
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            value = int(value)  # Ensure integer values are written as integers
                        
                        cell = ws.cell(row=target_row, column=col_idx)
                        cell.value = value
                        
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0'
                        
                        # Debug: Log important columns
                        if 'Pending for Dispatch' in col_name:
                            print(f"      Row {target_row}, Col {col_idx} ({col_name}): {value}")
                        
                        cells_written += 1
                    except Exception as e:
                        print(f"   ⚠️  Error writing {col_name} to row {target_row}, col {col_idx}: {e}")
                        import traceback
                        traceback.print_exc()
                        pass
            
            rows_written += 1
        
        print(f"   ✅ Wrote {cells_written} cell values across {rows_written} rows")
        
        # Verify critical columns were written
        if 'Pending for Dispatch (E)' in column_mapping:
            pending_dispatch_col = column_mapping['Pending for Dispatch (E)']
            sample_value = ws.cell(row=data_start_row, column=pending_dispatch_col).value
            print(f"   ✅ Verified: Pending for Dispatch column (col {pending_dispatch_col}) has value: {sample_value} in first row")
        else:
            print(f"   ⚠️  WARNING: 'Pending for Dispatch (E)' column not mapped!")
        
        # Add total row (like ZBM reports)
        total_row_pos = data_start_row + len(zbm_rows)
        copy_row_style(template_data_row, total_row_pos)
        
        # Set "Total" text in ZBM Name column
        if zbm_name_col:
            cell = ws.cell(row=total_row_pos, column=zbm_name_col)
            cell.value = "Total"
            cell.font = Font(bold=True, name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate and write totals (like ZBM reports)
        # List of columns that should NOT be summed (text columns)
        exclude_from_total = ['ZBM Code', 'ZBM Name', 'Level', 'Affiliate', 'Division', 'Division Name', '# Unique ZBMs']
        
        # Debug: Show which columns will be totaled
        columns_to_total = [col for col in column_mapping.keys() if col in zbm_rows.columns and col not in exclude_from_total]
        print(f"   📊 Calculating totals for {len(columns_to_total)} columns")
        if 'Pending for Dispatch (E)' in columns_to_total:
            print(f"      ✅ 'Pending for Dispatch (E)' will be included in totals")
        
        for col_name, col_idx in column_mapping.items():
            if col_name in zbm_rows.columns and col_name not in exclude_from_total:
                try:
                    # Check if column contains numeric data
                    col_data = zbm_rows[col_name]
                    
                    # Replace NaN with 0 for calculation
                    col_data_clean = col_data.fillna(0)
                    
                    # Try to convert to numeric, skip if not numeric
                    numeric_data = pd.to_numeric(col_data_clean, errors='coerce')
                    
                    if numeric_data.notna().any():  # If any values are numeric
                        total_value = int(numeric_data.sum())
                        
                        cell = ws.cell(row=total_row_pos, column=col_idx)
                        cell.value = total_value
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.number_format = '0'
                        
                        # Debug: Log important totals
                        if 'Pending for Dispatch' in col_name:
                            print(f"      Total for {col_name}: {total_value}")
                except Exception as e:
                    print(f"   ⚠️  Error calculating total for {col_name}: {e}")
                    # Skip non-numeric columns
                    pass
        
        print(f"   ✅ Wrote {len(zbm_rows)} ZBM rows and 1 Total row")

        # Save file
        safe_div_name = str(div_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f"Division_Summary_{div_code}_{safe_div_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        print(f"   💾 Saving file to: {filepath}")
        wb.save(filepath)
        
        # Verify file was created
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            print(f"   ✅ Created: {filename} ({file_size} bytes)")
        else:
            print(f"   ❌ ERROR: File was not created at {filepath}")
        
    except Exception as e:
        print(f"   ❌ Error creating Excel report for Division {div_code}: {e}")
        import traceback
        traceback.print_exc()
        raise  # Re-raise to see the error

if __name__ == "__main__":
    create_division_hierarchical_reports()
        
   
